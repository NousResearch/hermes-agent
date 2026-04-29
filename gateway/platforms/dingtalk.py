"""
DingTalk platform adapter using Stream Mode.

Uses dingtalk-stream SDK (>=0.20) for real-time message reception without webhooks.
Responses are sent via DingTalk's session webhook (markdown format).
Supports: text, images, audio, video, rich text, files, and group @mentions.

Requires:
    pip install "dingtalk-stream>=0.20" httpx
    DINGTALK_CLIENT_ID and DINGTALK_CLIENT_SECRET env vars

Configuration in config.yaml:
    platforms:
      dingtalk:
        enabled: true
        # Optional group-chat gating (mirrors Slack/Telegram/Discord):
        require_mention: true            # or DINGTALK_REQUIRE_MENTION env var
        # free_response_chats:           # conversations that skip require_mention
        #   - cidABC==
        # mention_patterns:              # regex wake-words (e.g. Chinese bot names)
        #   - "^小马"
        # allowed_users:                 # staff_id or sender_id list; "*" = any
        #   - "manager1234"
        extra:
          client_id: "your-app-key"      # or DINGTALK_CLIENT_ID env var
          client_secret: "your-secret"   # or DINGTALK_CLIENT_SECRET env var
"""

import asyncio
import json
import logging
import os
import re
import time
import traceback
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Set

try:
    import dingtalk_stream
    from dingtalk_stream import ChatbotMessage
    from dingtalk_stream.frames import CallbackMessage, AckMessage

    DINGTALK_STREAM_AVAILABLE = True
except ImportError:
    DINGTALK_STREAM_AVAILABLE = False
    dingtalk_stream = None  # type: ignore[assignment]
    ChatbotMessage = None  # type: ignore[assignment]
    CallbackMessage = None  # type: ignore[assignment]
    AckMessage = type(
        "AckMessage",
        (),
        {
            "STATUS_OK": 200,
            "STATUS_SYSTEM_EXCEPTION": 500,
        },
    )  # type: ignore[assignment]

try:
    import httpx

    HTTPX_AVAILABLE = True
except ImportError:
    HTTPX_AVAILABLE = False
    httpx = None  # type: ignore[assignment]

# Card SDK for AI Cards (following QwenPaw pattern)
try:
    from alibabacloud_dingtalk.card_1_0 import (
        client as dingtalk_card_client,
        models as dingtalk_card_models,
    )
    from alibabacloud_dingtalk.robot_1_0 import (
        client as dingtalk_robot_client,
        models as dingtalk_robot_models,
    )
    from alibabacloud_tea_openapi import models as open_api_models
    from alibabacloud_tea_util import models as tea_util_models

    CARD_SDK_AVAILABLE = True
except ImportError:
    CARD_SDK_AVAILABLE = False
    dingtalk_card_client = None
    dingtalk_card_models = None
    dingtalk_robot_client = None
    dingtalk_robot_models = None
    open_api_models = None
    tea_util_models = None

from gateway.config import Platform, PlatformConfig
from gateway.platforms.helpers import MessageDeduplicator
from gateway.platforms.base import (
    BasePlatformAdapter,
    MessageEvent,
    MessageType,
    SendResult,
)

logger = logging.getLogger(__name__)

MAX_MESSAGE_LENGTH = 20000
RECONNECT_BACKOFF = [2, 5, 10, 30, 60]
_SESSION_WEBHOOKS_MAX = 500
_DINGTALK_WEBHOOK_RE = re.compile(r'^https://(?:api|oapi)\.dingtalk\.com/')

# DingTalk message type → runtime content type
DINGTALK_TYPE_MAPPING = {
    "picture": "image",
    "voice": "audio",
}


_DINGTALK_OAUTH_TOKEN_URL = "https://api.dingtalk.com/v1.0/oauth2/accessToken"
_DINGTALK_OAPI_TOKEN_URL = "https://oapi.dingtalk.com/gettoken"  # legacy, for /media/upload
_DINGTALK_OAPI_MEDIA_UPLOAD_URL = "https://oapi.dingtalk.com/media/upload"
_DINGTALK_OTO_BATCH_SEND_URL = "https://api.dingtalk.com/v1.0/robot/oToMessages/batchSend"
_DINGTALK_GROUP_SEND_URL = "https://api.dingtalk.com/v1.0/robot/groupMessages/send"

# Process-wide access token cache: (client_id) -> (token, expires_at_ts).
# Tokens are valid for ~2h; we refresh with a 5-minute safety margin so parallel
# sends share one round-trip.  Two caches because /media/upload needs the
# legacy OAPI token (different endpoint, different token) — matches the
# reference connector's getAccessToken / getOapiAccessToken split.
_DINGTALK_TOKEN_CACHE: Dict[str, tuple[str, float]] = {}
_DINGTALK_OAPI_TOKEN_CACHE: Dict[str, tuple[str, float]] = {}
_DINGTALK_TOKEN_LOCK = asyncio.Lock()
_DINGTALK_OAPI_TOKEN_LOCK = asyncio.Lock()
_DINGTALK_TOKEN_SAFETY_MARGIN = 300.0  # refresh 5 min before expiry

# Media upload limits.  DingTalk /media/upload caps at 20 MB per file.
_DINGTALK_MEDIA_MAX_SIZE = 20 * 1024 * 1024
# File-extension → (media_type, msg_key) routing for /media/upload.
_IMAGE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"})
_VIDEO_EXTS = frozenset({".mp4", ".mov", ".avi", ".mkv", ".webm"})
_VOICE_EXTS = frozenset({".amr", ".mp3", ".wav", ".aac", ".m4a", ".ogg"})

# LWCP-encoded sender_id (opaque DingTalk-internal form).  Treated as OTO but
# the OpenAPI will reject it as ``staffId.notExisted`` unless the caller has
# resolved it to a real staffId via the contact APIs first.
_LWCP_SENDER_RE = re.compile(r'^\$:LWCP_')
# Real openConversationId (group chat): ``cid...==`` base64-padded form.
_OPEN_CONVERSATION_RE = re.compile(r'^cid[A-Za-z0-9+/_\-]+={0,2}$')

# DingTalk message type → runtime content type
DINGTALK_TYPE_MAPPING = {
    "picture": "image",
    "voice": "audio",
}

# ---------------------------------------------------------------------------
# File content auto-parsing (ported from dingtalk-openclaw-connector
# core/message-handler.ts:700-956)
# ---------------------------------------------------------------------------

# Extensions that can be parsed as plain text and injected into agent context.
_TEXT_FILE_EXTS = frozenset({
    ".txt", ".md", ".json", ".xml", ".yaml", ".yml", ".csv", ".log",
    ".js", ".ts", ".py", ".java", ".c", ".cpp", ".h", ".sh", ".bat",
    ".html", ".css", ".sql", ".rb", ".go", ".rs", ".toml", ".ini", ".cfg",
})
_DOCX_EXTS = frozenset({".docx", ".doc"})
_PDF_EXTS = frozenset({".pdf"})
_EXCEL_EXTS = frozenset({".xlsx", ".xls", ".xlsm"})
# All parseable extensions (text + docx + pdf + excel).
_PARSEABLE_FILE_EXTS = _TEXT_FILE_EXTS | _DOCX_EXTS | _PDF_EXTS | _EXCEL_EXTS


def _file_type_label(ext: str) -> str:
    """Return a human-readable Chinese label for a file extension."""
    if ext in _TEXT_FILE_EXTS:
        return "文本文件"
    if ext in _DOCX_EXTS:
        return "Word 文档"
    if ext in _PDF_EXTS:
        return "PDF 文档"
    if ext in {".xlsx", ".xls"}:
        return "Excel 表格"
    if ext in {".pptx", ".ppt"}:
        return "PPT 演示文稿"
    if ext in {".zip", ".rar", ".7z", ".tar", ".gz"}:
        return "压缩包"
    if ext in _IMAGE_EXTS:
        return "图片"
    if ext in _VIDEO_EXTS:
        return "视频"
    if ext in _VOICE_EXTS:
        return "音频"
    return "文件"


def _parse_text_file(file_path: str) -> Optional[str]:
    """Read a plain-text file and return its content."""
    try:
        text = Path(file_path).read_text(encoding="utf-8", errors="replace").strip()
        return text if text else None
    except Exception as exc:
        logger.warning("Failed to read text file %s: %s", file_path, exc)
        return None


def _parse_docx_file(file_path: str) -> Optional[str]:
    """Extract raw text from a .docx file using python-docx."""
    try:
        import docx  # python-docx
    except ImportError:
        logger.warning(
            "python-docx not installed, cannot parse .docx. "
            "Install with: pip install python-docx"
        )
        return None
    try:
        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs).strip()
        return text if text else None
    except Exception as exc:
        logger.warning("Failed to parse docx %s: %s", file_path, exc)
        return None


def _parse_pdf_file(file_path: str) -> Optional[str]:
    """Extract text from a PDF file.

    Tries pdfplumber first (better table / layout handling), then falls back
    to PyPDF2.
    """
    # Try pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages_text).strip()
        if text:
            return text
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("pdfplumber failed for %s: %s", file_path, exc)

    # Fallback to PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        pages_text = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages_text).strip()
        return text if text else None
    except ImportError:
        logger.warning(
            "Neither pdfplumber nor PyPDF2 installed, cannot parse PDF. "
            "Install with: pip install pdfplumber  or  pip install PyPDF2"
        )
        return None
    except Exception as exc:
        logger.warning("PyPDF2 failed for %s: %s", file_path, exc)
        return None


def _parse_excel_file(file_path: str) -> Optional[str]:
    """Extract text representation from an Excel (.xlsx/.xls/.xlsm) file.

    Tries openpyxl first (modern xlsx), then falls back to xlrd (legacy xls).
    Converts each sheet into a Markdown-style table so the LLM can reason
    over the data directly.
    """
    # -- Try openpyxl (xlsx / xlsm) --
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_text: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(cells)
            if not rows:
                continue
            # Build markdown table
            header = "| " + " | ".join(rows[0]) + " |"
            sep = "| " + " | ".join(["---"] * len(rows[0])) + " |"
            body_lines = []
            for r in rows[1:]:
                # Pad or truncate to header length
                padded = r + [""] * (len(rows[0]) - len(r))
                body_lines.append("| " + " | ".join(padded[:len(rows[0])]) + " |")
            table = "\n".join([header, sep] + body_lines)
            sheets_text.append(f"### Sheet: {sheet_name}\n\n{table}")
        wb.close()
        text = "\n\n".join(sheets_text).strip()
        if text:
            return text
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("openpyxl failed for %s: %s", file_path, exc)

    # -- Fallback: pandas (handles both xlsx and xls) --
    try:
        import pandas as pd
        xls = pd.ExcelFile(file_path)
        sheets_text = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            md_table = df.to_markdown(index=False)
            if md_table:
                sheets_text.append(f"### Sheet: {sheet_name}\n\n{md_table}")
        text = "\n\n".join(sheets_text).strip()
        return text if text else None
    except ImportError:
        logger.warning(
            "Neither openpyxl nor pandas installed, cannot parse Excel. "
            "Install with: pip install openpyxl  or  pip install pandas"
        )
        return None
    except Exception as exc:
        logger.warning("pandas Excel parse failed for %s: %s", file_path, exc)
        return None


def _parse_file_content(file_path: str, file_name: str) -> Optional[str]:
    """Dispatch to the correct parser based on file extension.

    Returns the extracted text, or None if unparseable / binary.
    """
    ext = os.path.splitext(file_name)[1].lower()
    if ext in _TEXT_FILE_EXTS:
        return _parse_text_file(file_path)
    if ext in _DOCX_EXTS:
        return _parse_docx_file(file_path)
    if ext in _PDF_EXTS:
        return _parse_pdf_file(file_path)
    if ext in _EXCEL_EXTS:
        return _parse_excel_file(file_path)
    return None


# Persistent inbound-media storage.  Files live here so agent tools
# (vision_analyze / transcribe_audio / read_file …) can reopen them after
# ``_on_message`` returns.  Cleanup on adapter connect ages out files
# older than 24 h so the directory doesn't grow unboundedly.

def _inbound_media_dir() -> str:
    base = os.environ.get("HERMES_HOME") or os.path.expanduser("~/.hermes")
    path = os.path.join(base, "inbound_media")
    os.makedirs(path, exist_ok=True)
    return path


def _cleanup_inbound_media(older_than_seconds: int = 24 * 3600) -> int:
    """Delete inbox files older than the cutoff.  Returns the count removed."""
    removed = 0
    try:
        dir_path = _inbound_media_dir()
        now = time.time()
        for name in os.listdir(dir_path):
            full = os.path.join(dir_path, name)
            try:
                if os.path.isfile(full) and (now - os.path.getmtime(full)) > older_than_seconds:
                    os.unlink(full)
                    removed += 1
            except OSError:
                continue
    except Exception:
        logger.debug("inbound media cleanup failed", exc_info=True)
    return removed


async def _download_file_to_inbox(
    url: str, file_name: str, *, msg_id: str = "", timeout: float = 60.0,
) -> Optional[str]:
    """Download *url* into the persistent inbox; agent tools can reopen it.

    ``msg_id`` is hashed into the output filename so concurrent messages
    don't clobber each other.
    """
    if not HTTPX_AVAILABLE:
        return None
    try:
        ext = os.path.splitext(file_name)[1] or ""
        base = re.sub(r'[^\w.-]', '_', os.path.splitext(file_name)[0])[:80] or "file"
        slug = re.sub(r'[^\w]', '', msg_id)[:16] if msg_id else uuid.uuid4().hex[:12]
        out_path = os.path.join(_inbound_media_dir(), f"{slug}_{base}{ext}")
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
        with open(out_path, "wb") as fh:
            fh.write(resp.content)
        logger.info(
            "Downloaded inbound media %s (%d bytes) -> %s",
            file_name, len(resp.content), out_path,
        )
        return out_path
    except Exception as exc:
        logger.warning("Failed to download inbound media %s: %s", file_name, exc)
        return None


# Extension → MIME type used when we hand a local inbound file off to the
# agent via ``event.media_urls`` / ``media_types``.
_EXT_TO_MIME = {
    ".mp3": "audio/mpeg", ".m4a": "audio/mp4", ".wav": "audio/wav",
    ".ogg": "audio/ogg", ".amr": "audio/amr", ".aac": "audio/aac",
    ".mp4": "video/mp4", ".mov": "video/quicktime", ".webm": "video/webm",
    ".avi": "video/x-msvideo", ".mkv": "video/x-matroska",
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp", ".svg": "image/svg+xml",
    ".pdf": "application/pdf", ".md": "text/markdown", ".txt": "text/plain",
    ".csv": "text/csv", ".json": "application/json",
    ".zip": "application/zip", ".tar": "application/x-tar", ".gz": "application/gzip",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _mime_for_file(file_name: str) -> str:
    ext = os.path.splitext(file_name)[1].lower()
    return _EXT_TO_MIME.get(ext, "application/octet-stream")


async def _dingtalk_fetch_access_token(
    client_id: str, client_secret: str
) -> str:
    """Return a cached DingTalk accessToken, refreshing if near expiry.

    Uses ``/v1.0/oauth2/accessToken`` directly (no SDK dependency) so the
    ``send_message`` tool can call this without opening a Stream WebSocket
    that would conflict with the running gateway's exclusive connection.
    """
    if not HTTPX_AVAILABLE:
        raise RuntimeError("httpx not installed")

    now = datetime.now(tz=timezone.utc).timestamp()
    cached = _DINGTALK_TOKEN_CACHE.get(client_id)
    if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
        return cached[0]

    async with _DINGTALK_TOKEN_LOCK:
        cached = _DINGTALK_TOKEN_CACHE.get(client_id)
        if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
            return cached[0]
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(
                _DINGTALK_OAUTH_TOKEN_URL,
                json={"appKey": client_id, "appSecret": client_secret},
            )
        resp.raise_for_status()
        body = resp.json()
        token = body.get("accessToken") or ""
        expires_in = body.get("expireIn") or body.get("expires_in") or 7200
        if not token:
            raise RuntimeError(
                f"DingTalk accessToken response missing token: {body}"
            )
        _DINGTALK_TOKEN_CACHE[client_id] = (
            token, now + float(expires_in),
        )
        return token


async def _dingtalk_fetch_oapi_token(
    client_id: str, client_secret: str,
) -> Optional[str]:
    """Fetch the legacy OAPI accessToken (for /media/upload).

    DingTalk has two parallel token systems:
      - ``/v1.0/oauth2/accessToken``  — new, for message OpenAPI
      - ``/gettoken``                 — legacy OAPI, for /media/upload
    Cached separately because they're different endpoints returning
    different tokens.  Returns None on any failure (non-fatal —
    callers should fall back to text-only).
    """
    if not HTTPX_AVAILABLE:
        return None
    now = datetime.now(tz=timezone.utc).timestamp()
    cached = _DINGTALK_OAPI_TOKEN_CACHE.get(client_id)
    if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
        return cached[0]

    async with _DINGTALK_OAPI_TOKEN_LOCK:
        cached = _DINGTALK_OAPI_TOKEN_CACHE.get(client_id)
        if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
            return cached[0]
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    _DINGTALK_OAPI_TOKEN_URL,
                    params={"appkey": client_id, "appsecret": client_secret},
                )
            body = resp.json()
            if body.get("errcode") != 0 or not body.get("access_token"):
                return None
            token = str(body["access_token"])
            expires_in = float(body.get("expires_in") or 7200)
            _DINGTALK_OAPI_TOKEN_CACHE[client_id] = (
                token, now + expires_in,
            )
            return token
        except Exception:
            return None


def _classify_media_kind(path: str) -> str:
    """Return one of 'image' / 'voice' / 'video' / 'file' for a local path."""
    ext = os.path.splitext(path)[1].lower()
    if ext in _IMAGE_EXTS:
        return "image"
    if ext in _VOICE_EXTS:
        return "voice"
    if ext in _VIDEO_EXTS:
        return "video"
    return "file"


async def _dingtalk_upload_media(
    *,
    oapi_token: str,
    file_path: str,
    media_kind: str,
) -> Optional[str]:
    """Upload a local file to DingTalk /media/upload.

    Returns the raw ``media_id`` on success (preserving the leading ``@``
    when present), or None on failure.

    Caller is responsible for shaping the id per msgKey:
      - ``sampleImageMsg.photoURL``   → ``@<raw>``  (raw mediaId, NOT the
        CDN download URL — the CDN URL renders as a white placeholder in
        the DingTalk client; reference messaging.ts:714-719)
      - ``sampleFile.mediaId``        → ``@<raw>``
      - ``sampleAudio.mediaId``       → ``@<raw>``
      - ``sampleVideo.videoMediaId``  → ``@<raw>``

    Matches the reference connector's ``uploadMediaToDingTalk``
    (media/common.ts:65).  Large-file chunked upload (>20MB) not yet
    ported — callers should pre-split or downscale.
    """
    if not HTTPX_AVAILABLE:
        return None
    if not os.path.exists(file_path):
        return None
    size = os.path.getsize(file_path)
    if size > _DINGTALK_MEDIA_MAX_SIZE:
        logger.warning(
            "[dingtalk] media file %s is %d bytes, exceeds 20MB limit",
            file_path, size,
        )
        return None

    ext = os.path.splitext(file_path)[1].lower()
    if media_kind == "image":
        content_type = "image/png" if ext == ".png" else "image/jpeg"
    elif media_kind == "video":
        content_type = "video/mp4" if ext == ".mp4" else "video/quicktime"
    elif media_kind == "voice":
        content_type = "audio/mpeg" if ext == ".mp3" else "audio/amr"
    else:
        content_type = "application/octet-stream"

    # DingTalk OAPI quirk: videos must be uploaded with type=file
    # (reference: media/common.ts:141).
    upload_type = "file" if media_kind == "video" else media_kind

    try:
        with open(file_path, "rb") as fh:
            files = {
                "media": (
                    os.path.basename(file_path),
                    fh,
                    content_type,
                ),
            }
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.post(
                    _DINGTALK_OAPI_MEDIA_UPLOAD_URL,
                    params={"access_token": oapi_token, "type": upload_type},
                    files=files,
                )
        body = resp.json()
        if body.get("errcode") != 0:
            logger.warning(
                "[dingtalk] /media/upload errcode=%s errmsg=%s",
                body.get("errcode"), body.get("errmsg"),
            )
            return None
        media_id = body.get("media_id") or ""
        return media_id or None
    except Exception as e:
        logger.warning("[dingtalk] /media/upload failed: %s", e)
        return None


def _dingtalk_classify_chat_id(chat_id: str) -> tuple[str, str]:
    """Classify a DingTalk chat_id into (bucket, resolved_id).

    Accepts the same target-prefix convention the dingtalk-openclaw-connector
    reference uses (``src/services/messaging.ts:sendTextToDingTalk``):

    - ``"group:<openConversationId>"`` → group
    - ``"user:<staffId>"``              → oto
    - ``"cid...=="``                    → group (auto-detect)
    - ``"$:LWCP_..."``                  → lwcp (fast-fail, needs contact perms)
    - anything else                     → oto (assume staffId)

    Returns ``(bucket, id)`` where ``id`` is the stripped value to put into
    the OpenAPI payload.
    """
    if not chat_id:
        return "oto", ""
    if chat_id.startswith("group:"):
        return "group", chat_id[6:]
    if chat_id.startswith("user:"):
        return "oto", chat_id[5:]
    if _LWCP_SENDER_RE.match(chat_id):
        return "lwcp", chat_id
    if _OPEN_CONVERSATION_RE.match(chat_id):
        return "group", chat_id
    return "oto", chat_id


# Smart markdown detection — content contains any of these characters/leading
# markers, assume markdown and upgrade msgKey to sampleMarkdown (reference:
# messaging.ts:820-826).  Plain-text prose uses sampleText so DingTalk's
# client renders it without extra padding/heading inference.
_MARKDOWN_LEADING_RE = re.compile(r'^[#*>\-]')
_MARKDOWN_INLINE_RE = re.compile(r'[*_`#\[\]]')


def _looks_like_markdown(content: str) -> bool:
    if not content:
        return False
    first_line = content.lstrip().split("\n", 1)[0]
    if _MARKDOWN_LEADING_RE.match(first_line):
        return True
    if _MARKDOWN_INLINE_RE.search(content):
        return True
    if "\n" in content:  # multi-line prose renders better as markdown
        return True
    return False


def _dingtalk_build_msg_param(
    content: str,
    *,
    title: str = "Hermes",
    msg_type: Optional[str] = None,
) -> tuple[str, str]:
    """Return (msgKey, JSON-encoded msgParam).

    ``msg_type`` overrides auto-detection:
      - "text"      → sampleText
      - "markdown"  → sampleMarkdown
      - "image"     → sampleImageMsg (content = mediaId or photoURL)
      - "file"      → sampleFile (content = mediaId)
      - "voice"     → sampleAudio (content = mediaId; duration=0 per reference)
      - "video"     → sampleVideo (content = mediaId)
    If None, chooses sampleMarkdown vs sampleText via
    ``_looks_like_markdown``.  Matches the reference's ``buildMsgPayload``
    (messaging.ts:150+).
    """
    truncated = content[:MAX_MESSAGE_LENGTH]

    if msg_type == "image":
        # sampleImageMsg.photoURL accepts the raw media_id with the leading
        # ``@`` preserved — DingTalk's server-side resolves this to the
        # tenant-internal image store.  Reference (messaging.ts:714-719 +
        # messaging.ts:188-192): `photoURL: uploadResult.mediaId` with the
        # inline comment "使用原始 mediaId（带 @）".
        #
        # Earlier iterations passed `https://down.dingtalk.com/media/<clean>`
        # here.  The API accepted it, but DingTalk clients rendered a blank
        # (white) image because the CDN URL requires a tenant-scoped auth
        # handshake that the client doesn't perform during `sampleImageMsg`
        # rendering.  Only keep the URL passthrough for already-public http(s)
        # URLs the caller provides directly.
        if truncated.startswith(("http://", "https://")):
            photo_url = truncated
        else:
            photo_url = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleImageMsg", json.dumps(
            {"photoURL": photo_url}, ensure_ascii=False,
        )
    if msg_type == "file":
        # sampleFile / sampleAudio / sampleVideo take the raw media_id with
        # leading ``@`` preserved (reference: media.ts:680+ & :876).
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleFile", json.dumps(
            {"mediaId": media_id, "fileName": title, "fileType": "file"},
            ensure_ascii=False,
        )
    if msg_type == "voice":
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleAudio", json.dumps(
            {"mediaId": media_id, "duration": "0"}, ensure_ascii=False,
        )
    if msg_type == "video":
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleVideo", json.dumps(
            {"videoMediaId": media_id, "videoType": "mp4",
             "picMediaId": "", "duration": "0"},
            ensure_ascii=False,
        )
    if msg_type == "text":
        return "sampleText", json.dumps(
            {"content": truncated}, ensure_ascii=False,
        )
    if msg_type == "markdown" or _looks_like_markdown(truncated):
        return "sampleMarkdown", json.dumps(
            {"title": title, "text": truncated}, ensure_ascii=False,
        )
    return "sampleText", json.dumps(
        {"content": truncated}, ensure_ascii=False,
    )


async def _dingtalk_post_one(
    *,
    token: str,
    robot_code: str,
    bucket: str,
    target_id: str,
    msg_key: str,
    msg_param: str,
) -> Dict[str, Any]:
    """POST a single already-built message to the correct Robot OpenAPI endpoint."""
    if bucket == "group":
        url = _DINGTALK_GROUP_SEND_URL
        payload = {
            "robotCode": robot_code,
            "openConversationId": target_id,
            "msgKey": msg_key,
            "msgParam": msg_param,
        }
    else:
        url = _DINGTALK_OTO_BATCH_SEND_URL
        payload = {
            "robotCode": robot_code,
            "userIds": [target_id],
            "msgKey": msg_key,
            "msgParam": msg_param,
        }

    headers = {
        "x-acs-dingtalk-access-token": token,
        "Content-Type": "application/json",
    }
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(url, json=payload, headers=headers)
    except Exception as e:
        return {"error": f"DingTalk OpenAPI request failed: {e}"}

    if resp.status_code >= 300:
        return {
            "error": f"DingTalk OpenAPI HTTP {resp.status_code}: {resp.text[:300]}"
        }

    try:
        body = resp.json()
    except Exception:
        body = {}
    # Robot OpenAPI returns ``processQueryKey`` on success (reference:
    # openclaw-connector messaging.ts:254,331,999).  Fallback to requestId.
    request_id = body.get("processQueryKey") or body.get("requestId") or ""
    return {"success": True, "request_id": request_id}


async def dingtalk_send_proactive(
    *,
    client_id: str,
    client_secret: str,
    robot_code: str,
    chat_id: str,
    content: str,
    media_files: Optional[List[tuple[str, bool]]] = None,
    title: str = "Hermes",
) -> Dict[str, Any]:
    """Send a message via DingTalk Robot OpenAPI without Stream Mode.

    Proactive-send counterpart to ``DingTalkAdapter.send()``'s
    session_webhook path.  Authenticates with AppKey/Secret, routes by
    ``chat_id`` shape (group / oto / lwcp), and optionally uploads+sends
    media attachments as additional messages.

    ``media_files`` is a list of ``(local_path, is_voice_hint)`` tuples.
    Each file gets uploaded to /media/upload, then sent as a dedicated
    message (sampleImageMsg / sampleAudio / sampleVideo / sampleFile).
    Text is sent first (if non-empty), then each media in order.  Mirrors
    Feishu's approach of emitting separate messages per attachment.

    ``title`` is used for sampleMarkdown headings when content auto-upgrades.

    Returns ``{"success": True, "request_id": ..., "route": ...}`` on success
    of the primary (text or first) message.  ``media_results`` lists each
    media attempt with its outcome.
    """
    if not HTTPX_AVAILABLE:
        return {"error": "httpx not installed"}
    if not client_id or not client_secret:
        return {"error": "DingTalk proactive send requires client_id + client_secret"}
    if not robot_code:
        return {"error": "DingTalk proactive send requires robot_code (defaults to client_id)"}
    if not chat_id:
        return {"error": "DingTalk proactive send requires chat_id"}

    # Classify FIRST so LWCP fast-fails with zero network roundtrip.
    bucket, target_id = _dingtalk_classify_chat_id(chat_id)
    if bucket == "lwcp":
        return {
            "error": (
                "DingTalk chat_id is LWCP-encoded sender_id (%s…): cannot be routed "
                "directly by Robot OpenAPI. Grant the app 'qyapi_get_department_list' "
                "+ 'qyapi_get_department_member' permissions at open-dev.dingtalk.com "
                "and resolve to plain staffId before calling proactive send."
            ) % target_id[:16]
        }

    try:
        token = await _dingtalk_fetch_access_token(client_id, client_secret)
    except Exception as e:
        return {"error": f"DingTalk accessToken fetch failed: {e}"}

    results: List[Dict[str, Any]] = []
    primary_request_id = ""

    # 1. Text first (if non-empty).  Smart markdown detection applied here.
    if content and content.strip():
        msg_key, msg_param = _dingtalk_build_msg_param(content, title=title)
        text_result = await _dingtalk_post_one(
            token=token, robot_code=robot_code, bucket=bucket,
            target_id=target_id, msg_key=msg_key, msg_param=msg_param,
        )
        if text_result.get("error"):
            return text_result
        primary_request_id = text_result.get("request_id", "")

    # 2. Media (one message per file).  Non-fatal per file — we collect
    # per-file results and let the caller decide what to surface.
    if media_files:
        oapi_token = await _dingtalk_fetch_oapi_token(client_id, client_secret)
        if not oapi_token:
            results.append({
                "error": (
                    "Could not obtain OAPI token for /media/upload — verify the "
                    "app has ServerAPI permissions (qyapi_media) and retry."
                ),
            })
            # Keep the text result if we had one.
            if primary_request_id:
                return {
                    "success": True,
                    "request_id": primary_request_id,
                    "chat_id": target_id,
                    "route": bucket,
                    "media_results": results,
                }
            return results[-1]

        for raw_path, is_voice_hint in media_files:
            media_kind = (
                "voice" if is_voice_hint and _classify_media_kind(raw_path) in ("voice", "file")
                else _classify_media_kind(raw_path)
            )
            media_id = await _dingtalk_upload_media(
                oapi_token=oapi_token,
                file_path=raw_path,
                media_kind=media_kind,
            )
            if not media_id:
                results.append({
                    "error": f"upload failed for {os.path.basename(raw_path)}",
                    "path": raw_path,
                })
                continue
            m_msg_key, m_msg_param = _dingtalk_build_msg_param(
                media_id,
                msg_type=media_kind,
                title=os.path.basename(raw_path),
            )
            m_result = await _dingtalk_post_one(
                token=token, robot_code=robot_code, bucket=bucket,
                target_id=target_id, msg_key=m_msg_key, msg_param=m_msg_param,
            )
            m_result["kind"] = media_kind
            m_result["path"] = raw_path
            results.append(m_result)
            if not primary_request_id and m_result.get("success"):
                primary_request_id = m_result.get("request_id", "")

    if not primary_request_id and not content.strip():
        # Nothing delivered (empty text, all media failed).
        return {
            "error": "Nothing delivered: empty text and all media uploads failed",
            "media_results": results,
        }

    return {
        "success": True,
        "request_id": primary_request_id,
        "chat_id": target_id,
        "route": bucket,
        **({"media_results": results} if results else {}),
    }


def check_dingtalk_requirements() -> bool:
    """Check if DingTalk dependencies are available and configured."""
    if not DINGTALK_STREAM_AVAILABLE or not HTTPX_AVAILABLE:
        return False
    if not os.getenv("DINGTALK_CLIENT_ID") or not os.getenv("DINGTALK_CLIENT_SECRET"):
        return False
    return True


class DingTalkAdapter(BasePlatformAdapter):
    """DingTalk chatbot adapter using Stream Mode.

    The dingtalk-stream SDK maintains a long-lived WebSocket connection.
    Incoming messages arrive via a ChatbotHandler callback. Replies are
    sent via the incoming message's session_webhook URL using httpx.

    Features:
    - Text messages (plain + rich text)
    - Images, audio, video, files (via download codes)
    - Group chat @mention detection
    - Session webhook caching with expiry tracking
    - Markdown formatted replies
    """

    MAX_MESSAGE_LENGTH = MAX_MESSAGE_LENGTH

    @property
    def SUPPORTS_MESSAGE_EDITING(self) -> bool:  # noqa: N802
        """Edits only meaningful when AI Cards are configured.

        The gateway gates streaming cursor + edit behaviour on this flag,
        so we must reflect the actual adapter capability at runtime.
        """
        return bool(self._card_template_id and self._card_sdk)

    @property
    def REQUIRES_EDIT_FINALIZE(self) -> bool:  # noqa: N802
        """AI Card lifecycle requires an explicit ``finalize=True`` edit
        to close the streaming indicator, even when the final content is
        identical to the last streamed update.  Enabled only when cards
        are configured — webhook-only DingTalk doesn't need it.
        """
        return bool(self._card_template_id and self._card_sdk)

    def __init__(self, config: PlatformConfig):
        super().__init__(config, Platform.DINGTALK)

        extra = config.extra or {}
        self._client_id: str = extra.get("client_id") or os.getenv(
            "DINGTALK_CLIENT_ID", ""
        )
        self._client_secret: str = extra.get("client_secret") or os.getenv(
            "DINGTALK_CLIENT_SECRET", ""
        )

        # Group-chat gating (mirrors Slack/Telegram/Discord/WhatsApp conventions).
        # Mention state is the structured ``is_in_at_list`` attribute from the
        # dingtalk-stream SDK (set from the callback's ``isInAtList`` flag),
        # not text parsing.
        self._mention_patterns: List[re.Pattern] = self._compile_mention_patterns()
        self._allowed_users: Set[str] = self._load_allowed_users()

        self._stream_client: Any = None
        self._stream_task: Optional[asyncio.Task] = None
        self._http_client: Optional["httpx.AsyncClient"] = None
        self._card_sdk: Optional[Any] = None
        self._robot_sdk: Optional[Any] = None
        self._robot_code: str = extra.get("robot_code") or self._client_id

        # Message deduplication
        self._dedup = MessageDeduplicator(max_size=1000)
        # Map chat_id -> (session_webhook, expired_time_ms) for reply routing
        self._session_webhooks: Dict[str, tuple[str, int]] = {}
        # Map chat_id -> last inbound ChatbotMessage. Keyed by chat_id instead
        # of a single class attribute to avoid cross-message clobbering when
        # multiple conversations run concurrently.
        self._message_contexts: Dict[str, Any] = {}
        self._card_template_id: Optional[str] = extra.get("card_template_id")

        # Chats for which we've already fired the Done reaction — prevents
        # double-firing across segment boundaries or parallel flows
        # (tool-progress + stream-consumer both finalizing their cards).
        # Reset each inbound message.
        self._done_emoji_fired: Set[str] = set()
        # Cards in streaming state per chat: chat_id -> { out_track_id -> last_content }.
        # Every `send()` creates+finalizes a card (closed state).  A subsequent
        # `edit_message(finalize=False)` re-opens the card (DingTalk's API
        # allows streaming_update on a finalized card — it flips back to
        # streaming).  We track those reopened cards so the next `send()` can
        # auto-close them as siblings — otherwise tool-progress cards get
        # stuck in streaming state forever.
        self._streaming_cards: Dict[str, Dict[str, str]] = {}
        # Track fire-and-forget emoji/reaction coroutines so Python's GC
        # doesn't drop them mid-flight, and we can cancel them on disconnect.
        self._bg_tasks: Set[asyncio.Task] = set()

    # -- Connection lifecycle -----------------------------------------------

    async def connect(self) -> bool:
        """Connect to DingTalk via Stream Mode."""
        if not DINGTALK_STREAM_AVAILABLE:
            logger.warning(
                "[%s] dingtalk-stream not installed. Run: pip install 'dingtalk-stream>=0.20'",
                self.name,
            )
            return False
        if not HTTPX_AVAILABLE:
            logger.warning(
                "[%s] httpx not installed. Run: pip install httpx", self.name
            )
            return False
        if not self._client_id or not self._client_secret:
            logger.warning(
                "[%s] DINGTALK_CLIENT_ID and DINGTALK_CLIENT_SECRET required", self.name
            )
            return False

        try:
            self._http_client = httpx.AsyncClient(timeout=30.0)

            credential = dingtalk_stream.Credential(
                self._client_id, self._client_secret
            )
            self._stream_client = dingtalk_stream.DingTalkStreamClient(credential)

            # Initialize card SDK if available and configured
            if CARD_SDK_AVAILABLE and self._card_template_id:
                sdk_config = open_api_models.Config()
                sdk_config.protocol = "https"
                sdk_config.region_id = "central"
                self._card_sdk = dingtalk_card_client.Client(sdk_config)
                self._robot_sdk = dingtalk_robot_client.Client(sdk_config)
                logger.info(
                    "[%s] Card SDK initialized with template: %s",
                    self.name,
                    self._card_template_id,
                )
            elif CARD_SDK_AVAILABLE:
                # Initialize robot SDK even without card template (for media download)
                sdk_config = open_api_models.Config()
                sdk_config.protocol = "https"
                sdk_config.region_id = "central"
                self._robot_sdk = dingtalk_robot_client.Client(sdk_config)
                logger.info("[%s] Robot SDK initialized (media download)", self.name)

            # Capture the current event loop for cross-thread dispatch
            loop = asyncio.get_running_loop()
            handler = _IncomingHandler(self, loop)
            self._stream_client.register_callback_handler(
                dingtalk_stream.ChatbotMessage.TOPIC, handler
            )

            self._stream_task = asyncio.create_task(self._run_stream())
            self._mark_connected()
            logger.info("[%s] Connected via Stream Mode", self.name)
            return True
        except Exception as e:
            logger.error("[%s] Failed to connect: %s", self.name, e)
            return False

    async def _run_stream(self) -> None:
        """Run the async stream client with auto-reconnection."""
        backoff_idx = 0
        while self._running:
            try:
                logger.debug("[%s] Starting stream client...", self.name)
                await self._stream_client.start()
            except asyncio.CancelledError:
                return
            except Exception as e:
                if not self._running:
                    return
                logger.warning("[%s] Stream client error: %s", self.name, e)

            if not self._running:
                return

            delay = RECONNECT_BACKOFF[min(backoff_idx, len(RECONNECT_BACKOFF) - 1)]
            logger.info("[%s] Reconnecting in %ds...", self.name, delay)
            await asyncio.sleep(delay)
            backoff_idx += 1

    async def disconnect(self) -> None:
        """Disconnect from DingTalk."""
        self._running = False
        self._mark_disconnected()

        # Close the active websocket first so the stream task sees the
        # disconnection and exits cleanly, rather than getting stuck
        # awaiting frames that will never arrive.
        websocket = getattr(self._stream_client, "websocket", None) if self._stream_client else None
        if websocket is not None:
            try:
                await websocket.close()
            except Exception as e:
                logger.debug("[%s] websocket close during disconnect failed: %s", self.name, e)

        if self._stream_task:
            # Try graceful close first if SDK supports it. The SDK's close()
            # is sync and may block on network I/O, so offload to a thread.
            if hasattr(self._stream_client, "close"):
                try:
                    await asyncio.to_thread(self._stream_client.close)
                except Exception:
                    pass

            self._stream_task.cancel()
            try:
                await asyncio.wait_for(self._stream_task, timeout=5.0)
            except (asyncio.CancelledError, asyncio.TimeoutError):
                logger.debug("[%s] stream task did not exit cleanly during disconnect", self.name)
            self._stream_task = None

        # Cancel any in-flight background tasks (emoji reactions, etc.)
        if self._bg_tasks:
            for task in list(self._bg_tasks):
                task.cancel()
            await asyncio.gather(*self._bg_tasks, return_exceptions=True)
            self._bg_tasks.clear()

        if self._http_client:
            await self._http_client.aclose()
            self._http_client = None

        self._stream_client = None
        self._session_webhooks.clear()
        self._message_contexts.clear()
        self._streaming_cards.clear()
        self._done_emoji_fired.clear()
        self._dedup.clear()
        logger.info("[%s] Disconnected", self.name)

    # -- Group gating --------------------------------------------------------

    def _dingtalk_require_mention(self) -> bool:
        """Return whether group chats should require an explicit bot trigger."""
        configured = self.config.extra.get("require_mention")
        if configured is not None:
            if isinstance(configured, str):
                return configured.lower() in ("true", "1", "yes", "on")
            return bool(configured)
        return os.getenv("DINGTALK_REQUIRE_MENTION", "false").lower() in ("true", "1", "yes", "on")

    def _dingtalk_free_response_chats(self) -> Set[str]:
        raw = self.config.extra.get("free_response_chats")
        if raw is None:
            raw = os.getenv("DINGTALK_FREE_RESPONSE_CHATS", "")
        if isinstance(raw, list):
            return {str(part).strip() for part in raw if str(part).strip()}
        return {part.strip() for part in str(raw).split(",") if part.strip()}

    def _compile_mention_patterns(self) -> List[re.Pattern]:
        """Compile optional regex wake-word patterns for group triggers."""
        patterns = self.config.extra.get("mention_patterns") if self.config.extra else None
        if patterns is None:
            raw = os.getenv("DINGTALK_MENTION_PATTERNS", "").strip()
            if raw:
                try:
                    loaded = json.loads(raw)
                except Exception:
                    loaded = [part.strip() for part in raw.splitlines() if part.strip()]
                    if not loaded:
                        loaded = [part.strip() for part in raw.split(",") if part.strip()]
                patterns = loaded

        if patterns is None:
            return []
        if isinstance(patterns, str):
            patterns = [patterns]
        if not isinstance(patterns, list):
            logger.warning(
                "[%s] dingtalk mention_patterns must be a list or string; got %s",
                self.name,
                type(patterns).__name__,
            )
            return []

        compiled: List[re.Pattern] = []
        for pattern in patterns:
            if not isinstance(pattern, str) or not pattern.strip():
                continue
            try:
                compiled.append(re.compile(pattern, re.IGNORECASE))
            except re.error as exc:
                logger.warning("[%s] Invalid DingTalk mention pattern %r: %s", self.name, pattern, exc)
        if compiled:
            logger.info("[%s] Loaded %d DingTalk mention pattern(s)", self.name, len(compiled))
        return compiled

    def _load_allowed_users(self) -> Set[str]:
        """Load allowed-users list from config.extra or env var.

        IDs are matched case-insensitively against the sender's ``staff_id`` and
        ``sender_id``. A wildcard ``*`` disables the check.
        """
        raw = self.config.extra.get("allowed_users") if self.config.extra else None
        if raw is None:
            raw = os.getenv("DINGTALK_ALLOWED_USERS", "")
        if isinstance(raw, list):
            items = [str(part).strip() for part in raw if str(part).strip()]
        else:
            items = [part.strip() for part in str(raw).split(",") if part.strip()]
        return {item.lower() for item in items}

    def _is_user_allowed(self, sender_id: str, sender_staff_id: str) -> bool:
        if not self._allowed_users or "*" in self._allowed_users:
            return True
        candidates = {(sender_id or "").lower(), (sender_staff_id or "").lower()}
        candidates.discard("")
        return bool(candidates & self._allowed_users)

    def _message_mentions_bot(self, message: "ChatbotMessage") -> bool:
        """True if the bot was @-mentioned in a group message.

        dingtalk-stream sets ``is_in_at_list`` on the incoming ChatbotMessage
        when the bot is addressed via @-mention.
        """
        return bool(getattr(message, "is_in_at_list", False))

    def _message_matches_mention_patterns(self, text: str) -> bool:
        if not text or not self._mention_patterns:
            return False
        return any(pattern.search(text) for pattern in self._mention_patterns)

    def _should_process_message(self, message: "ChatbotMessage", text: str, is_group: bool, chat_id: str) -> bool:
        """Apply DingTalk group trigger rules.

        DMs remain unrestricted (subject to ``allowed_users`` which is enforced
        earlier). Group messages are accepted when:
        - the chat is explicitly allowlisted in ``free_response_chats``
        - ``require_mention`` is disabled
        - the bot is @mentioned (``is_in_at_list``)
        - the text matches a configured regex wake-word pattern
        """
        if not is_group:
            return True
        if chat_id and chat_id in self._dingtalk_free_response_chats():
            return True
        if not self._dingtalk_require_mention():
            return True
        if self._message_mentions_bot(message):
            return True
        return self._message_matches_mention_patterns(text)

    def _spawn_bg(self, coro) -> None:
        """Start a fire-and-forget coroutine and track it for cleanup."""
        task = asyncio.create_task(coro)
        self._bg_tasks.add(task)
        task.add_done_callback(self._bg_tasks.discard)

    # -- AI Card lifecycle helpers ------------------------------------------

    async def _close_streaming_siblings(self, chat_id: str) -> None:
        """Finalize any previously-open streaming cards for this chat.

        Called at the start of every ``send()`` so lingering tool-progress
        cards that were reopened by ``edit_message(finalize=False)`` get
        cleanly closed before the next card is created.  Without this,
        tool-progress cards stay stuck in streaming state after the agent
        moves on (there is no explicit "turn end" signal from the gateway).
        """
        cards = self._streaming_cards.pop(chat_id, None)
        if not cards:
            return
        token = await self._get_access_token()
        if not token:
            return
        for out_track_id, last_content in list(cards.items()):
            try:
                await self._stream_card_content(
                    out_track_id, token, last_content, finalize=True,
                )
                logger.debug(
                    "[%s] AI Card sibling closed: %s",
                    self.name, out_track_id,
                )
            except Exception as e:
                logger.debug(
                    "[%s] Sibling close failed for %s: %s",
                    self.name, out_track_id, e,
                )

    def _fire_done_reaction(self, chat_id: str) -> None:
        """Swap 🤔Thinking → 🥳Done on the original user message.

        Idempotent per chat_id — safe to call from segment-break flushes
        and final-done flushes without double-firing.
        """
        if chat_id in self._done_emoji_fired:
            return
        self._done_emoji_fired.add(chat_id)
        msg = self._message_contexts.get(chat_id)
        if not msg:
            return
        msg_id = getattr(msg, "message_id", "") or ""
        conversation_id = getattr(msg, "conversation_id", "") or ""
        if not (msg_id and conversation_id):
            return

        async def _swap() -> None:
            await self._send_emotion(
                msg_id, conversation_id, "🤔Thinking", recall=True,
            )
            await self._send_emotion(
                msg_id, conversation_id, "🥳Done", recall=False,
            )

        self._spawn_bg(_swap())

    # -- Inbound message processing -----------------------------------------

    async def _on_message(
        self,
        message: "ChatbotMessage",
    ) -> None:
        """Process an incoming DingTalk chatbot message."""
        msg_id = getattr(message, "message_id", None) or uuid.uuid4().hex
        if self._dedup.is_duplicate(msg_id):
            logger.debug("[%s] Duplicate message %s, skipping", self.name, msg_id)
            return

        # Chat context
        conversation_id = getattr(message, "conversation_id", "") or ""
        conversation_type = getattr(message, "conversation_type", "1")
        is_group = str(conversation_type) == "2"
        sender_id = getattr(message, "sender_id", "") or ""
        sender_nick = getattr(message, "sender_nick", "") or sender_id
        sender_staff_id = getattr(message, "sender_staff_id", "") or ""

        chat_id = conversation_id or sender_id
        chat_type = "group" if is_group else "dm"

        # Allowed-users gate (applies to both DM and group)
        if not self._is_user_allowed(sender_id, sender_staff_id):
            logger.debug(
                "[%s] Dropping message from non-allowlisted user staff_id=%s sender_id=%s",
                self.name, sender_staff_id, sender_id,
            )
            return

        # Group mention/pattern gate.  DMs pass through unconditionally.
        # We need the message text for regex wake-word matching; extract it
        # early but don't consume the rest of the pipeline until after the
        # gate decides whether to process.
        _early_text = self._extract_text(message) or ""
        if not self._should_process_message(message, _early_text, is_group, chat_id):
            logger.debug(
                "[%s] Dropping group message that failed mention gate message_id=%s chat_id=%s",
                self.name, msg_id, chat_id,
            )
            return

        # Stash the incoming message keyed by chat_id so concurrent
        # conversations don't clobber each other's context.  Also reset
        # the per-chat "Done emoji fired" marker so a new inbound message
        # gets its own Thinking→Done cycle.
        if chat_id:
            self._message_contexts[chat_id] = message
            self._done_emoji_fired.discard(chat_id)

        # Store session webhook
        session_webhook = getattr(message, "session_webhook", None) or ""
        session_webhook_expired_time = (
            getattr(message, "session_webhook_expired_time", 0) or 0
        )
        if session_webhook and chat_id and _DINGTALK_WEBHOOK_RE.match(session_webhook):
            if len(self._session_webhooks) >= _SESSION_WEBHOOKS_MAX:
                try:
                    self._session_webhooks.pop(next(iter(self._session_webhooks)))
                except StopIteration:
                    pass
            self._session_webhooks[chat_id] = (
                session_webhook,
                session_webhook_expired_time,
            )

        # Resolve media download codes to URLs so vision tools can use them
        await self._resolve_media_codes(message)

        # Extract text content
        text = self._extract_text(message)

        # Determine message type and build media list
        msg_type, media_urls, media_types = self._extract_media(message)

        if not text and not media_urls:
            logger.debug("[%s] Empty message, skipping", self.name)
            return

        source = self.build_source(
            chat_id=chat_id,
            chat_name=getattr(message, "conversation_title", None),
            chat_type=chat_type,
            user_id=sender_id,
            user_name=sender_nick,
            user_id_alt=sender_staff_id if sender_staff_id else None,
        )

        # Parse timestamp
        create_at = getattr(message, "create_at", None)
        try:
            timestamp = (
                datetime.fromtimestamp(int(create_at) / 1000, tz=timezone.utc)
                if create_at
                else datetime.now(tz=timezone.utc)
            )
        except (ValueError, OSError, TypeError):
            timestamp = datetime.now(tz=timezone.utc)

        event = MessageEvent(
            text=text,
            message_type=msg_type,
            source=source,
            message_id=msg_id,
            raw_message=message,
            media_urls=media_urls,
            media_types=media_types,
            timestamp=timestamp,
        )

        logger.debug(
            "[%s] Message from %s in %s: %s",
            self.name,
            sender_nick,
            chat_id[:20] if chat_id else "?",
            text[:80] if text else "(media)",
        )
        await self.handle_message(event)

    @staticmethod
    def _extract_text(message: "ChatbotMessage") -> str:
        """Extract plain text from a DingTalk chatbot message.

        Handles both legacy and current dingtalk-stream SDK payload shapes:
          * legacy: ``message.text`` was a dict ``{"content": "..."}``
          * >= 0.20: ``message.text`` is a ``TextContent`` dataclass whose
            ``__str__`` returns ``"TextContent(content=...)"`` — never fall
            back to ``str(text)`` without extracting ``.content`` first.
          * rich text moved from ``message.rich_text`` (list) to
            ``message.rich_text_content.rich_text_list`` (list of dicts).
        """
        text = getattr(message, "text", None) or ""

        # Handle TextContent object (SDK style)
        if hasattr(text, "content"):
            content = (text.content or "").strip()
        elif isinstance(text, dict):
            content = text.get("content", "").strip()
        else:
            content = str(text).strip()

        if not content:
            rich_text = getattr(message, "rich_text_content", None) or getattr(
                message, "rich_text", None
            )
            if rich_text:
                rich_list = getattr(rich_text, "rich_text_list", None) or rich_text
                if isinstance(rich_list, list):
                    parts = []
                    for item in rich_list:
                        if isinstance(item, dict):
                            t = item.get("text") or item.get("content") or ""
                            if t:
                                parts.append(t)
                        elif hasattr(item, "text") and item.text:
                            parts.append(item.text)
                    content = " ".join(parts).strip()

        # Do NOT strip "@bot" from the text.  The mention is a routing
        # signal (delivered structurally via callback `isInAtList`), and
        # regex-stripping @handles would collateral-damage e-mails
        # (alice@example.com), SSH URLs (git@github.com), and literal
        # references the user wrote ("what does @openai think").  Let the
        # LLM see the raw text — it handles "@bot hello" cleanly.
        return content

    def _extract_media(self, message: "ChatbotMessage"):
        """Extract media info from message. Returns (MessageType, [urls], [mime_types])."""
        msg_type = MessageType.TEXT
        media_urls = []
        media_types = []

        # Check for image/picture
        image_content = getattr(message, "image_content", None)
        if image_content:
            download_code = getattr(image_content, "download_code", None)
            if download_code:
                media_urls.append(download_code)
                media_types.append("image")
                msg_type = MessageType.PHOTO

        # Check for rich text with mixed content
        rich_text = getattr(message, "rich_text_content", None) or getattr(
            message, "rich_text", None
        )
        if rich_text:
            rich_list = getattr(rich_text, "rich_text_list", None) or rich_text
            if isinstance(rich_list, list):
                for item in rich_list:
                    if isinstance(item, dict):
                        dl_code = (
                            item.get("downloadCode") or item.get("download_code") or ""
                        )
                        item_type = item.get("type", "")
                        if dl_code:
                            mapped = DINGTALK_TYPE_MAPPING.get(item_type, "file")
                            media_urls.append(dl_code)
                            if mapped == "image":
                                media_types.append("image")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.PHOTO
                            elif mapped == "audio":
                                media_types.append("audio")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.AUDIO
                            elif mapped == "video":
                                media_types.append("video")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.VIDEO
                            else:
                                media_types.append("application/octet-stream")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.DOCUMENT

        msg_type_str = getattr(message, "message_type", "") or ""
        if msg_type_str == "picture" and not media_urls:
            msg_type = MessageType.PHOTO
        elif msg_type_str == "richText":
            msg_type = (
                MessageType.PHOTO
                if any("image" in t for t in media_types)
                else MessageType.TEXT
            )

        return msg_type, media_urls, media_types

    # -- Outbound messaging -------------------------------------------------

    async def send(
        self,
        chat_id: str,
        content: str,
        reply_to: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> SendResult:
        """Send a markdown reply via DingTalk session webhook."""
        metadata = metadata or {}
        logger.debug(
            "[%s] send() chat_id=%s card_enabled=%s",
            self.name,
            chat_id,
            bool(self._card_template_id and self._card_sdk),
        )

        # Check metadata first (for direct webhook sends)
        session_webhook = metadata.get("session_webhook")
        if not session_webhook:
            webhook_info = self._get_valid_webhook(chat_id)
            if webhook_info:
                session_webhook, _ = webhook_info
            else:
                # No reply webhook for this chat — this is a proactive send
                # (no preceding inbound message in cache).  Fall back to
                # Robot OpenAPI using AppKey/Secret, same pattern Feishu uses
                # for proactive ``im.v1.message.create`` calls.
                logger.debug(
                    "[%s] No session_webhook for chat_id=%s; routing via Robot OpenAPI",
                    self.name, chat_id,
                )
                result = await dingtalk_send_proactive(
                    client_id=self._client_id,
                    client_secret=self._client_secret,
                    robot_code=self._robot_code,
                    chat_id=chat_id,
                    content=content,
                )
                if result.get("success"):
                    return SendResult(
                        success=True,
                        message_id=result.get("request_id") or uuid.uuid4().hex[:12],
                    )
                return SendResult(
                    success=False, error=result.get("error", "unknown"),
                )

        if not self._http_client:
            return SendResult(success=False, error="HTTP client not initialized")

        # Look up the inbound message for this chat (for AI Card routing)
        current_message = self._message_contexts.get(chat_id)

        # ``reply_to`` is the signal that this send is the FINAL response
        # to an inbound user message — only `base.py:_send_with_retry` sets
        # it.  Tool-progress, commentary, and stream-consumer first-sends
        # all leave it None.  We use it for two orthogonal decisions:
        #   1. finalize on create?  Yes if final reply, No if intermediate
        #      (intermediate cards stay in streaming state so edit_message
        #      updates don't flicker closed→streaming→closed repeatedly).
        #   2. fire Done reaction?  Only when this is the final reply.
        is_final_reply = reply_to is not None

        # Try AI Card first (using alibabacloud_dingtalk.card_1_0 SDK).
        if self._card_template_id and current_message and self._card_sdk:
            # Close any previously-open streaming cards for this chat
            # before creating a new one (handles tool-progress → final-
            # response handoff; also cleans up lingering commentary cards).
            await self._close_streaming_siblings(chat_id)

            result = await self._create_and_stream_card(
                chat_id, current_message, content,
                finalize=is_final_reply,
            )
            if result and result.success:
                if is_final_reply:
                    # Final reply: card closed, swap Thinking → Done.
                    self._fire_done_reaction(chat_id)
                else:
                    # Intermediate (tool progress / commentary / streaming
                    # first chunk): keep the card open and track it so the
                    # next send() auto-closes it as a sibling, or
                    # edit_message(finalize=True) closes it explicitly.
                    self._streaming_cards.setdefault(chat_id, {})[
                        result.message_id
                    ] = content
                return result

            logger.warning("[%s] AI Card send failed, falling back to webhook", self.name)

        logger.debug("[%s] Sending via webhook", self.name)
        # Normalize markdown for DingTalk
        normalized = self._normalize_markdown(content[: self.MAX_MESSAGE_LENGTH])

        payload = {
            "msgtype": "markdown",
            "markdown": {"title": "Hermes", "text": normalized},
        }

        try:
            resp = await self._http_client.post(
                session_webhook, json=payload, timeout=15.0
            )
            if resp.status_code < 300:
                # Webhook path: fire Done only for final replies, same as
                # the card path.
                if is_final_reply:
                    self._fire_done_reaction(chat_id)
                return SendResult(success=True, message_id=uuid.uuid4().hex[:12])
            body = resp.text
            logger.warning(
                "[%s] Send failed HTTP %d: %s", self.name, resp.status_code, body[:200]
            )
            return SendResult(
                success=False, error=f"HTTP {resp.status_code}: {body[:200]}"
            )
        except httpx.TimeoutException:
            return SendResult(
                success=False, error="Timeout sending message to DingTalk"
            )
        except Exception as e:
            logger.error("[%s] Send error: %s", self.name, e)
            return SendResult(success=False, error=str(e))

    async def send_typing(self, chat_id: str, metadata=None) -> None:
        """DingTalk does not support typing indicators."""
        pass

    async def get_chat_info(self, chat_id: str) -> Dict[str, Any]:
        """Return basic info about a DingTalk conversation."""
        return {
            "name": chat_id,
            "type": "group" if "group" in chat_id.lower() else "dm",
        }

    def _get_valid_webhook(self, chat_id: str) -> Optional[tuple[str, int]]:
        """Get a valid (non-expired) session webhook for the given chat_id."""
        info = self._session_webhooks.get(chat_id)
        if not info:
            return None
        webhook, expired_time_ms = info
        # Check expiry with 5-minute safety margin
        if expired_time_ms and expired_time_ms > 0:
            now_ms = int(datetime.now(tz=timezone.utc).timestamp() * 1000)
            safety_margin_ms = 5 * 60 * 1000
            if now_ms + safety_margin_ms >= expired_time_ms:
                # Expired, remove from cache
                self._session_webhooks.pop(chat_id, None)
                return None
        return info

    async def _create_and_stream_card(
        self,
        chat_id: str,
        message: Any,
        content: str,
        *,
        finalize: bool = True,
    ) -> Optional[SendResult]:
        """Create an AI Card, deliver it to the conversation, and stream initial content.

        Always called with ``finalize=True`` from ``send()`` (closed state).
        If the caller later issues ``edit_message(finalize=False)``, the
        DingTalk streaming_update API reopens the card into streaming
        state, and we track that in ``_streaming_cards`` for sibling
        cleanup on the next send.
        """
        try:
            token = await self._get_access_token()
            if not token:
                return None

            out_track_id = f"hermes_{uuid.uuid4().hex[:12]}"

            conversation_id = getattr(message, "conversation_id", "") or ""
            conversation_type = getattr(message, "conversation_type", "1")
            is_group = str(conversation_type) == "2"
            sender_staff_id = getattr(message, "sender_staff_id", "") or ""

            runtime = tea_util_models.RuntimeOptions()

            # Step 1: Create card with STREAM callback type
            create_request = dingtalk_card_models.CreateCardRequest(
                card_template_id=self._card_template_id,
                out_track_id=out_track_id,
                card_data=dingtalk_card_models.CreateCardRequestCardData(
                    card_param_map={"content": ""},
                ),
                callback_type="STREAM",
                im_group_open_space_model=(
                    dingtalk_card_models.CreateCardRequestImGroupOpenSpaceModel(
                        support_forward=True,
                    )
                ),
                im_robot_open_space_model=(
                    dingtalk_card_models.CreateCardRequestImRobotOpenSpaceModel(
                        support_forward=True,
                    )
                ),
            )

            create_headers = dingtalk_card_models.CreateCardHeaders(
                x_acs_dingtalk_access_token=token,
            )

            await self._card_sdk.create_card_with_options_async(
                create_request, create_headers, runtime
            )

            # Step 2: Deliver card to the conversation
            if is_group:
                open_space_id = f"dtv1.card//IM_GROUP.{conversation_id}"
                deliver_request = dingtalk_card_models.DeliverCardRequest(
                    out_track_id=out_track_id,
                    user_id_type=1,
                    open_space_id=open_space_id,
                    im_group_open_deliver_model=(
                        dingtalk_card_models.DeliverCardRequestImGroupOpenDeliverModel(
                            robot_code=self._robot_code,
                        )
                    ),
                )
            else:
                if not sender_staff_id:
                    logger.warning(
                        "[%s] AI Card skipped: missing sender_staff_id for DM",
                        self.name,
                    )
                    return None
                open_space_id = f"dtv1.card//IM_ROBOT.{sender_staff_id}"
                deliver_request = dingtalk_card_models.DeliverCardRequest(
                    out_track_id=out_track_id,
                    user_id_type=1,
                    open_space_id=open_space_id,
                    im_robot_open_deliver_model=(
                        dingtalk_card_models.DeliverCardRequestImRobotOpenDeliverModel(
                            space_type="IM_ROBOT",
                        )
                    ),
                )

            deliver_headers = dingtalk_card_models.DeliverCardHeaders(
                x_acs_dingtalk_access_token=token,
            )

            await self._card_sdk.deliver_card_with_options_async(
                deliver_request, deliver_headers, runtime
            )

            # Step 3: Stream initial content.  finalize=True closes the
            # card immediately (one-shot); finalize=False keeps it open
            # for streaming edit_message updates by out_track_id.
            await self._stream_card_content(
                out_track_id, token, content, finalize=finalize,
            )

            logger.info(
                "[%s] AI Card %s: %s",
                self.name,
                "created+finalized" if finalize else "created (streaming)",
                out_track_id,
            )
            return SendResult(success=True, message_id=out_track_id)

        except Exception as e:
            logger.warning(
                "[%s] AI Card create failed: %s\n%s",
                self.name, e, traceback.format_exc(),
            )
            return None

    async def edit_message(
        self,
        chat_id: str,
        message_id: str,
        content: str,
        *,
        finalize: bool = False,
    ) -> SendResult:
        """Edit an AI Card by streaming updated content.

        ``message_id`` is the out_track_id returned by the initial ``send()``
        call that created this card.  Callers (stream_consumer, tool
        progress) track their own ids independently so two parallel flows
        on the same chat_id don't interfere.
        """
        if not message_id:
            return SendResult(success=False, error="message_id required")
        token = await self._get_access_token()
        if not token:
            return SendResult(success=False, error="No access token")

        try:
            await self._stream_card_content(
                message_id, token, content, finalize=finalize,
            )
            if finalize:
                # Remove from streaming-cards tracking and fire Done.  This
                # is the canonical "response ended" signal from stream
                # consumer's final edit.
                self._streaming_cards.get(chat_id, {}).pop(message_id, None)
                if not self._streaming_cards.get(chat_id):
                    self._streaming_cards.pop(chat_id, None)
                logger.debug(
                    "[%s] AI Card finalized (edit): %s",
                    self.name, message_id,
                )
                self._fire_done_reaction(chat_id)
            else:
                # Non-final edit reopens the card into streaming state —
                # track it so the next send() can auto-close it as a
                # sibling.
                self._streaming_cards.setdefault(chat_id, {})[message_id] = content
            return SendResult(success=True, message_id=message_id)
        except Exception as e:
            logger.warning("[%s] Card edit failed: %s", self.name, e)
            return SendResult(success=False, error=str(e))

    async def _stream_card_content(
        self,
        out_track_id: str,
        token: str,
        content: str,
        finalize: bool = False,
    ) -> None:
        """Stream content to an existing AI Card."""
        stream_request = dingtalk_card_models.StreamingUpdateRequest(
            out_track_id=out_track_id,
            guid=str(uuid.uuid4()),
            key="content",
            content=content[: self.MAX_MESSAGE_LENGTH],
            is_full=True,
            is_finalize=finalize,
            is_error=False,
        )

        stream_headers = dingtalk_card_models.StreamingUpdateHeaders(
            x_acs_dingtalk_access_token=token,
        )

        runtime = tea_util_models.RuntimeOptions()
        await self._card_sdk.streaming_update_with_options_async(
            stream_request, stream_headers, runtime
        )

    async def _get_access_token(self) -> Optional[str]:
        """Get access token using SDK's cached token."""
        if not self._stream_client:
            return None
        try:
            # SDK's get_access_token is sync and uses requests
            token = await asyncio.to_thread(self._stream_client.get_access_token)
            return token
        except Exception as e:
            logger.error("[%s] Failed to get access token: %s", self.name, e)
            return None

    async def _send_emotion(
        self,
        open_msg_id: str,
        open_conversation_id: str,
        emoji_name: str,
        *,
        recall: bool = False,
    ) -> None:
        """Add or recall an emoji reaction on a message."""
        if not self._robot_sdk or not open_msg_id or not open_conversation_id:
            return
        action = "recall" if recall else "reply"
        try:
            token = await self._get_access_token()
            if not token:
                return

            emotion_kwargs = {
                "robot_code": self._robot_code,
                "open_msg_id": open_msg_id,
                "open_conversation_id": open_conversation_id,
                "emotion_type": 2,
                "emotion_name": emoji_name,
            }
            runtime = tea_util_models.RuntimeOptions()

            if recall:
                emotion_kwargs["text_emotion"] = (
                    dingtalk_robot_models.RobotRecallEmotionRequestTextEmotion(
                        emotion_id="2659900",
                        emotion_name=emoji_name,
                        text=emoji_name,
                        background_id="im_bg_1",
                    )
                )
                request = dingtalk_robot_models.RobotRecallEmotionRequest(
                    **emotion_kwargs,
                )
                sdk_headers = dingtalk_robot_models.RobotRecallEmotionHeaders(
                    x_acs_dingtalk_access_token=token,
                )
                await self._robot_sdk.robot_recall_emotion_with_options_async(
                    request, sdk_headers, runtime
                )
            else:
                emotion_kwargs["text_emotion"] = (
                    dingtalk_robot_models.RobotReplyEmotionRequestTextEmotion(
                        emotion_id="2659900",
                        emotion_name=emoji_name,
                        text=emoji_name,
                        background_id="im_bg_1",
                    )
                )
                request = dingtalk_robot_models.RobotReplyEmotionRequest(
                    **emotion_kwargs,
                )
                sdk_headers = dingtalk_robot_models.RobotReplyEmotionHeaders(
                    x_acs_dingtalk_access_token=token,
                )
                await self._robot_sdk.robot_reply_emotion_with_options_async(
                    request, sdk_headers, runtime
                )
            logger.info(
                "[%s] _send_emotion: %s %s on msg=%s",
                self.name, action, emoji_name, open_msg_id[:24],
            )
        except Exception:
            logger.debug(
                "[%s] _send_emotion %s failed", self.name, action, exc_info=True
            )

    async def _resolve_media_codes(self, message: "ChatbotMessage") -> None:
        """Resolve download codes in message to actual URLs."""
        token = await self._get_access_token()
        if not token:
            return

        robot_code = getattr(message, "robot_code", None) or self._client_id
        codes_to_resolve = []

        # Collect codes and references to update
        # 1. Single image content
        img_content = getattr(message, "image_content", None)
        if img_content and getattr(img_content, "download_code", None):
            codes_to_resolve.append((img_content, "download_code"))

        # 2. Rich text list
        rich_text = getattr(message, "rich_text_content", None)
        if rich_text:
            rich_list = getattr(rich_text, "rich_text_list", []) or []
            for item in rich_list:
                if isinstance(item, dict):
                    for key in ("downloadCode", "pictureDownloadCode", "download_code"):
                        if item.get(key):
                            codes_to_resolve.append((item, key))

        if not codes_to_resolve:
            return

        # Resolve all codes in parallel
        tasks = []
        for obj, key in codes_to_resolve:
            code = getattr(obj, key, None) if hasattr(obj, key) else obj.get(key)
            if code:
                tasks.append(
                    self._fetch_download_url(code, robot_code, token, obj, key)
                )

        await asyncio.gather(*tasks, return_exceptions=True)

    async def _fetch_download_url(
        self, code: str, robot_code: str, token: str, obj, key: str
    ) -> None:
        """Fetch download URL for a single code using the robot SDK."""
        if not self._robot_sdk:
            logger.warning(
                "[%s] Robot SDK not initialized, cannot resolve media code",
                self.name,
            )
            return
        try:
            request = dingtalk_robot_models.RobotMessageFileDownloadRequest(
                download_code=code,
                robot_code=robot_code,
            )
            headers = dingtalk_robot_models.RobotMessageFileDownloadHeaders(
                x_acs_dingtalk_access_token=token,
            )
            runtime = tea_util_models.RuntimeOptions()
            response = await self._robot_sdk.robot_message_file_download_with_options_async(
                request, headers, runtime
            )
            body = response.body if response else None
            if body:
                url = getattr(body, "download_url", None)
                if url:
                    if hasattr(obj, key):
                        setattr(obj, key, url)
                    elif isinstance(obj, dict):
                        obj[key] = url
            else:
                logger.warning(
                    "[%s] Failed to download media: empty response for code %s",
                    self.name,
                    code,
                )
        except Exception as e:
            logger.error("[%s] Error resolving media code %s: %s", self.name, code, e)

    @staticmethod
    def _normalize_markdown(text: str) -> str:
        """Normalize markdown for DingTalk's parser.

        DingTalk's markdown renderer has quirks:
        - Numbered lists need blank line before them
        - Indented code blocks may render incorrectly
        """
        lines = text.split("\n")
        out = []
        for i, line in enumerate(lines):
            # Ensure blank line before numbered list items
            is_numbered = re.match(r"^\d+\.\s", line.strip())
            if is_numbered and i > 0:
                prev = lines[i - 1]
                if prev.strip() and not re.match(r"^\d+\.\s", prev.strip()):
                    out.append("")
            # Dedent fenced code blocks
            if line.strip().startswith("```") and line != line.lstrip():
                indent = len(line) - len(line.lstrip())
                line = line[indent:]
            out.append(line)
        return "\n".join(out)


# ---------------------------------------------------------------------------
# Internal stream handler
# ---------------------------------------------------------------------------


class _IncomingHandler(
    dingtalk_stream.ChatbotHandler if DINGTALK_STREAM_AVAILABLE else object
):
    """dingtalk-stream ChatbotHandler that forwards messages to the adapter.

    SDK >= 0.20 changed process() from sync to async, and the message
    parameter from ChatbotMessage to CallbackMessage. We parse the
    CallbackMessage.data dict into a ChatbotMessage before forwarding.
    """

    def __init__(self, adapter: DingTalkAdapter, loop: Optional[asyncio.AbstractEventLoop] = None):
        if DINGTALK_STREAM_AVAILABLE:
            super().__init__()
        self._adapter = adapter
        self._loop = loop

    async def process(self, message: "CallbackMessage"):
        """Called by dingtalk-stream (>=0.20) when a message arrives.

        dingtalk-stream >= 0.24 passes a CallbackMessage whose ``.data`` contains
        the chatbot payload. Convert it to ChatbotMessage via
        ``ChatbotMessage.from_dict()``.

        Message processing is dispatched as a background task so that this
        method returns the ACK immediately — blocking here would prevent the
        SDK from sending heartbeats, eventually causing a disconnect.
        """
        try:
            # CallbackMessage.data is a dict containing the raw DingTalk payload
            data = message.data
            if isinstance(data, str):
                data = json.loads(data)

            # Parse dict into ChatbotMessage using SDK's from_dict
            chatbot_msg = ChatbotMessage.from_dict(data)

            # Ensure session_webhook is populated even if the SDK's
            # from_dict() did not map it (field name mismatch across
            # SDK versions).
            if not getattr(chatbot_msg, "session_webhook", None):
                webhook = (
                    data.get("sessionWebhook")
                    or data.get("session_webhook")
                    or ""
                ) if isinstance(data, dict) else ""
                if webhook:
                    chatbot_msg.session_webhook = webhook

            # Ensure is_in_at_list is populated from the structured callback
            # flag even if from_dict() did not map it.  DingTalk sends
            # ``isInAtList`` in the raw payload; the adapter's mention check
            # reads the ChatbotMessage attribute ``is_in_at_list``.
            if not getattr(chatbot_msg, "is_in_at_list", False):
                raw_flag = (
                    data.get("isInAtList") if isinstance(data, dict) else False
                )
                if raw_flag:
                    chatbot_msg.is_in_at_list = True

            msg_id = getattr(chatbot_msg, "message_id", None) or ""
            conversation_id = getattr(chatbot_msg, "conversation_id", None) or ""

            # Thinking reaction — fire-and-forget, tracked
            if msg_id and conversation_id:
                self._adapter._spawn_bg(
                    self._adapter._send_emotion(
                        msg_id, conversation_id, "🤔Thinking", recall=False,
                    )
                )

            # Fire-and-forget: return ACK immediately, process in background.
            # Blocking here would prevent the SDK from sending heartbeats,
            # eventually causing a disconnect.  _on_message is wrapped so
            # exceptions inside the task surface in logs instead of
            # disappearing into the event loop.
            asyncio.create_task(self._safe_on_message(chatbot_msg))
        except Exception:
            logger.exception(
                "[%s] Error preparing incoming message", self._adapter.name
            )
            return AckMessage.STATUS_SYSTEM_EXCEPTION, "error"

        return AckMessage.STATUS_OK, "OK"

    async def _safe_on_message(self, chatbot_msg: "ChatbotMessage") -> None:
        """Wrapper that catches exceptions from _on_message."""
        try:
            await self._adapter._on_message(chatbot_msg)
        except Exception:
            logger.exception(
                "[%s] Error processing incoming message", self._adapter.name
            )
