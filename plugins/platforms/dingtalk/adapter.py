"""
DingTalk platform adapter using Stream Mode.

Uses dingtalk-stream SDK (>=0.20) for real-time message reception without webhooks.
Responses are sent via DingTalk's session webhook (markdown format).
Supports: text, images, audio, video, rich text, files, and group @mentions.

Requires:
    pip install "dingtalk-stream>=0.20" httpx
    DINGTALK_CLIENT_ID and DINGTALK_CLIENT_SECRET env vars

Configuration in config.yaml:
    platforms:
      dingtalk:
        enabled: true
        # Optional group-chat gating (mirrors Slack/Telegram/Discord):
        require_mention: true            # or DINGTALK_REQUIRE_MENTION env var
        # free_response_chats:           # conversations that skip require_mention
        #   - cidABC==
        # mention_patterns:              # regex wake-words (e.g. Chinese bot names)
        #   - "^小马"
        # allowed_users:                 # staff_id or sender_id list; "*" = any
        #   - "manager1234"
        extra:
          client_id: "your-app-key"      # or DINGTALK_CLIENT_ID env var
          client_secret: "your-secret"   # or DINGTALK_CLIENT_SECRET env var
"""

import asyncio
import json
import logging
import os
import random
import re
import time
import traceback
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import dingtalk_stream
    from dingtalk_stream import ChatbotMessage
    from dingtalk_stream.frames import CallbackMessage, AckMessage

    DINGTALK_STREAM_AVAILABLE = True
except Exception:  # noqa: BLE001 — broad: optional SDK's transitive deps (cryptography) may raise non-ImportError; degrade gracefully (#41112)
    DINGTALK_STREAM_AVAILABLE = False
    dingtalk_stream = None  # type: ignore[assignment]
    ChatbotMessage = None  # type: ignore[assignment]
    CallbackMessage = None  # type: ignore[assignment]
    AckMessage = type(
        "AckMessage",
        (),
        {
            "STATUS_OK": 200,
            "STATUS_SYSTEM_EXCEPTION": 500,
        },
    )  # type: ignore[assignment]

# Capture pristine ``websockets.connect`` at our import time so dingtalk-stream
# always uses the original even if another module (e.g. the Feishu adapter)
# later monkey-patches the global ``websockets.connect`` with an ``async def``
# wrapper — that turns the return value into a coroutine and breaks the SDK's
# ``async with websockets.connect(uri)`` at dingtalk_stream/stream.py:74.
try:
    import websockets as _pristine_ws_module

    _PRISTINE_WEBSOCKETS_CONNECT = _pristine_ws_module.connect
except ImportError:
    _pristine_ws_module = None  # type: ignore[assignment]
    _PRISTINE_WEBSOCKETS_CONNECT = None  # type: ignore[assignment]

try:
    import httpx

    HTTPX_AVAILABLE = True
except ImportError:
    HTTPX_AVAILABLE = False
    httpx = None  # type: ignore[assignment]


_WS_PROXY_INSTALLED = False


def _install_dingtalk_websockets_proxy() -> None:
    """Point dingtalk_stream.stream.websockets at a namespace holding the
    pristine ``connect`` captured at our import time.

    Idempotent. Safe no-op when dingtalk-stream or websockets is unavailable.
    """
    global _WS_PROXY_INSTALLED
    if _WS_PROXY_INSTALLED:
        return
    if _PRISTINE_WEBSOCKETS_CONNECT is None or dingtalk_stream is None:
        return
    try:
        import types

        from dingtalk_stream import stream as _dts_stream

        _dts_stream.websockets = types.SimpleNamespace(
            connect=_PRISTINE_WEBSOCKETS_CONNECT,
            exceptions=_pristine_ws_module.exceptions,
        )
        _WS_PROXY_INSTALLED = True
    except Exception:  # pragma: no cover - defensive
        logger.debug(
            "[DingTalk] failed to install websockets proxy on dingtalk_stream.stream",
            exc_info=True,
        )

# Card SDK for AI Cards (following QwenPaw pattern).
# Catch broad Exception, not just ImportError: the alibabacloud_dingtalk SDK
# transitively imports cryptography and can raise AttributeError (not
# ImportError) when the installed cryptography version skews from what the SDK
# expects (e.g. `cryptography.utils.DeprecatedIn46` missing on older
# cryptography). An optional SDK with a broken dependency chain must degrade
# gracefully — same as a missing one — rather than crash the whole adapter
# (and therefore the whole plugin) import. #41112.
try:
    from alibabacloud_dingtalk.card_1_0 import (
        client as dingtalk_card_client,
        models as dingtalk_card_models,
    )
    from alibabacloud_dingtalk.robot_1_0 import (
        client as dingtalk_robot_client,
        models as dingtalk_robot_models,
    )
    from alibabacloud_tea_openapi import models as open_api_models
    from alibabacloud_tea_util import models as tea_util_models

    CARD_SDK_AVAILABLE = True
except Exception:
    CARD_SDK_AVAILABLE = False
    dingtalk_card_client = None
    dingtalk_card_models = None
    dingtalk_robot_client = None
    dingtalk_robot_models = None
    open_api_models = None
    tea_util_models = None

from gateway.config import Platform, PlatformConfig
from gateway.platforms.helpers import MessageDeduplicator
from gateway.platforms.base import (
    BasePlatformAdapter,
    MessageEvent,
    MessageType,
    SendResult,
)

logger = logging.getLogger(__name__)

MAX_MESSAGE_LENGTH = 20000
RECONNECT_BACKOFF = [2, 5, 10, 30, 60]
_SESSION_WEBHOOKS_MAX = 500
_DINGTALK_WEBHOOK_RE = re.compile(r'^https://(?:api|oapi)\.dingtalk\.com/')

# AI Card streaming_update QPS guard.  The DingTalk gateway returns HTTP 403
# "concurrent update" when multiple edit_message calls hit the same card
# within ~500ms.  Reference (openclaw-connector reply-dispatcher.ts:103) uses
# 800ms as the per-card minimum interval between non-finalize edits.  We
# match that to avoid the same 403 storm.  Finalize edits are NEVER
# throttled — dropping them would leave the card stuck in streaming state.
_CARD_EDIT_THROTTLE_MS = 800
# Error-send cooldown so repeated failures don't spam users
# (reply-dispatcher.ts:108 uses 60s — same pattern, same reason).
_ERROR_COOLDOWN_MS = 60_000

# Global token-bucket rate for AI Card streaming_update across ALL chats.
# DingTalk's official cap is ~40 QPS per tenant; reference connector
# (messaging/card.ts:18) uses 20 as a safety margin.  Matching that, so
# concurrent sessions never bust the tenant-wide ceiling.
_CARD_API_MAX_QPS = 20
_CARD_API_QPS_BACKOFF_MS = 2_000

# Inbound-message queue TTL.  queueKey entries that haven't received a new
# message for this long are eligible for sweep (reference
# core/message-handler.ts:92 uses 5 min).
_INBOUND_QUEUE_TTL_SEC = 300
# Busy-ACK phrases when inbound queue already has a pending task.  Picked
# randomly so repeats don't feel scripted (reference utils/constants.ts
# QUEUE_BUSY_ACK_PHRASES).
_QUEUE_BUSY_ACK_PHRASES = (
    "收到，让我先把前一条处理完 🙏",
    "稍等，排队中……",
    "收到～手头这条完事就来",
    "别急，按顺序处理中",
)


class _CardTokenBucket:
    """Global async token bucket for DingTalk card streaming_update.

    Mirrors messaging/card.ts:23-95.  All streamAICard/edit_message calls
    across every chat + account share one bucket so concurrent sessions
    don't blow past the tenant-wide QPS limit and trigger 403 storms.
    Refills at ``rate`` tokens/second with capacity = rate.  On an
    upstream 403 limit, callers call ``trigger_backoff`` and subsequent
    acquirers wait out the backoff window.
    """

    def __init__(self, rate: float) -> None:
        self._rate = float(rate)
        self._tokens = float(rate)
        self._last_refill = time.monotonic()
        self._backoff_until = 0.0
        self._lock = asyncio.Lock()

    async def acquire(self) -> None:
        async with self._lock:
            now = time.monotonic()
            if now < self._backoff_until:
                await asyncio.sleep(self._backoff_until - now)
                now = time.monotonic()
            elapsed = now - self._last_refill
            self._tokens = min(self._rate, self._tokens + elapsed * self._rate)
            self._last_refill = now
            if self._tokens < 1.0:
                wait_s = (1.0 - self._tokens) / self._rate
                await asyncio.sleep(wait_s)
                self._tokens = 0.0
                self._last_refill = time.monotonic()
            else:
                self._tokens -= 1.0

    def trigger_backoff(self) -> None:
        self._backoff_until = time.monotonic() + _CARD_API_QPS_BACKOFF_MS / 1000.0


# Process-wide card-API rate limiter (shared across adapters).
_CARD_BUCKET = _CardTokenBucket(_CARD_API_MAX_QPS)

# DingTalk OpenAPI endpoints for proactive (non-session-webhook) messaging.
_DINGTALK_OAUTH_TOKEN_URL = "https://api.dingtalk.com/v1.0/oauth2/accessToken"
_DINGTALK_OAPI_TOKEN_URL = "https://oapi.dingtalk.com/gettoken"  # legacy, for /media/upload
_DINGTALK_OAPI_MEDIA_UPLOAD_URL = "https://oapi.dingtalk.com/media/upload"
_DINGTALK_OTO_BATCH_SEND_URL = "https://api.dingtalk.com/v1.0/robot/oToMessages/batchSend"
_DINGTALK_GROUP_SEND_URL = "https://api.dingtalk.com/v1.0/robot/groupMessages/send"

# Process-wide access token cache: (client_id) -> (token, expires_at_ts).
# Tokens are valid for ~2h; we refresh with a 5-minute safety margin so parallel
# sends share one round-trip.  Two caches because /media/upload needs the
# legacy OAPI token (different endpoint, different token) — matches the
# reference connector's getAccessToken / getOapiAccessToken split.
_DINGTALK_TOKEN_CACHE: Dict[str, tuple[str, float]] = {}
_DINGTALK_OAPI_TOKEN_CACHE: Dict[str, tuple[str, float]] = {}
_DINGTALK_TOKEN_LOCK = asyncio.Lock()
_DINGTALK_OAPI_TOKEN_LOCK = asyncio.Lock()
_DINGTALK_TOKEN_SAFETY_MARGIN = 300.0  # refresh 5 min before expiry

# Media upload limits.  DingTalk /media/upload caps at 20 MB per file.
_DINGTALK_MEDIA_MAX_SIZE = 20 * 1024 * 1024
# File-extension → (media_type, msg_key) routing for /media/upload.
_IMAGE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"})
_VIDEO_EXTS = frozenset({".mp4", ".mov", ".avi", ".mkv", ".webm"})
_VOICE_EXTS = frozenset({".amr", ".mp3", ".wav", ".aac", ".m4a", ".ogg"})

# LWCP-encoded sender_id (opaque DingTalk-internal form).  Treated as OTO but
# the OpenAPI will reject it as ``staffId.notExisted`` unless the caller has
# resolved it to a real staffId via the contact APIs first.
_LWCP_SENDER_RE = re.compile(r'^\$:LWCP_')
# Real openConversationId (group chat): ``cid...==`` base64-padded form.
_OPEN_CONVERSATION_RE = re.compile(r'^cid[A-Za-z0-9+/_\-]+={0,2}$')

# DingTalk message type → runtime content type
DINGTALK_TYPE_MAPPING = {
    "picture": "image",
    "voice": "audio",
}

# ---------------------------------------------------------------------------
# File content auto-parsing (ported from dingtalk-openclaw-connector
# core/message-handler.ts:700-956)
# ---------------------------------------------------------------------------

# Extensions that can be parsed as plain text and injected into agent context.
_TEXT_FILE_EXTS = frozenset({
    ".txt", ".md", ".json", ".xml", ".yaml", ".yml", ".csv", ".log",
    ".js", ".ts", ".py", ".java", ".c", ".cpp", ".h", ".sh", ".bat",
    ".html", ".css", ".sql", ".rb", ".go", ".rs", ".toml", ".ini", ".cfg",
})
_DOCX_EXTS = frozenset({".docx", ".doc"})
_PDF_EXTS = frozenset({".pdf"})
_EXCEL_EXTS = frozenset({".xlsx", ".xls", ".xlsm"})
# All parseable extensions (text + docx + pdf + excel).
_PARSEABLE_FILE_EXTS = _TEXT_FILE_EXTS | _DOCX_EXTS | _PDF_EXTS | _EXCEL_EXTS


def _file_type_label(ext: str) -> str:
    """Return a human-readable Chinese label for a file extension."""
    if ext in _TEXT_FILE_EXTS:
        return "文本文件"
    if ext in _DOCX_EXTS:
        return "Word 文档"
    if ext in _PDF_EXTS:
        return "PDF 文档"
    if ext in {".xlsx", ".xls"}:
        return "Excel 表格"
    if ext in {".pptx", ".ppt"}:
        return "PPT 演示文稿"
    if ext in {".zip", ".rar", ".7z", ".tar", ".gz"}:
        return "压缩包"
    if ext in _IMAGE_EXTS:
        return "图片"
    if ext in _VIDEO_EXTS:
        return "视频"
    if ext in _VOICE_EXTS:
        return "音频"
    return "文件"


def _parse_text_file(file_path: str) -> Optional[str]:
    """Read a plain-text file and return its content."""
    try:
        text = Path(file_path).read_text(encoding="utf-8", errors="replace").strip()
        return text if text else None
    except Exception as exc:
        logger.warning("Failed to read text file %s: %s", file_path, exc)
        return None


def _parse_docx_file(file_path: str) -> Optional[str]:
    """Extract raw text from a .docx file using python-docx."""
    try:
        import docx  # python-docx
    except ImportError:
        logger.warning(
            "python-docx not installed, cannot parse .docx. "
            "Install with: pip install python-docx"
        )
        return None
    try:
        doc = docx.Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs).strip()
        return text if text else None
    except Exception as exc:
        logger.warning("Failed to parse docx %s: %s", file_path, exc)
        return None


def _parse_pdf_file(file_path: str) -> Optional[str]:
    """Extract text from a PDF file.

    Tries pdfplumber first (better table / layout handling), then falls back
    to PyPDF2.
    """
    # Try pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            pages_text = [p.extract_text() or "" for p in pdf.pages]
        text = "\n".join(pages_text).strip()
        if text:
            return text
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("pdfplumber failed for %s: %s", file_path, exc)

    # Fallback to PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        pages_text = [page.extract_text() or "" for page in reader.pages]
        text = "\n".join(pages_text).strip()
        return text if text else None
    except ImportError:
        logger.warning(
            "Neither pdfplumber nor PyPDF2 installed, cannot parse PDF. "
            "Install with: pip install pdfplumber  or  pip install PyPDF2"
        )
        return None
    except Exception as exc:
        logger.warning("PyPDF2 failed for %s: %s", file_path, exc)
        return None


def _parse_excel_file(file_path: str) -> Optional[str]:
    """Extract text representation from an Excel (.xlsx/.xls/.xlsm) file.

    Tries openpyxl first (modern xlsx), then falls back to xlrd (legacy xls).
    Converts each sheet into a Markdown-style table so the LLM can reason
    over the data directly.
    """
    # -- Try openpyxl (xlsx / xlsm) --
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets_text: List[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(cells)
            if not rows:
                continue
            # Build markdown table
            header = "| " + " | ".join(rows[0]) + " |"
            sep = "| " + " | ".join(["---"] * len(rows[0])) + " |"
            body_lines = []
            for r in rows[1:]:
                # Pad or truncate to header length
                padded = r + [""] * (len(rows[0]) - len(r))
                body_lines.append("| " + " | ".join(padded[:len(rows[0])]) + " |")
            table = "\n".join([header, sep] + body_lines)
            sheets_text.append(f"### Sheet: {sheet_name}\n\n{table}")
        wb.close()
        text = "\n\n".join(sheets_text).strip()
        if text:
            return text
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("openpyxl failed for %s: %s", file_path, exc)

    # -- Fallback: pandas (handles both xlsx and xls) --
    try:
        import pandas as pd
        xls = pd.ExcelFile(file_path)
        sheets_text = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            md_table = df.to_markdown(index=False)
            if md_table:
                sheets_text.append(f"### Sheet: {sheet_name}\n\n{md_table}")
        text = "\n\n".join(sheets_text).strip()
        return text if text else None
    except ImportError:
        logger.warning(
            "Neither openpyxl nor pandas installed, cannot parse Excel. "
            "Install with: pip install openpyxl  or  pip install pandas"
        )
        return None
    except Exception as exc:
        logger.warning("pandas Excel parse failed for %s: %s", file_path, exc)
        return None


def _parse_file_content(file_path: str, file_name: str) -> Optional[str]:
    """Dispatch to the correct parser based on file extension.

    Returns the extracted text, or None if unparseable / binary.
    """
    ext = os.path.splitext(file_name)[1].lower()
    if ext in _TEXT_FILE_EXTS:
        return _parse_text_file(file_path)
    if ext in _DOCX_EXTS:
        return _parse_docx_file(file_path)
    if ext in _PDF_EXTS:
        return _parse_pdf_file(file_path)
    if ext in _EXCEL_EXTS:
        return _parse_excel_file(file_path)
    return None


# Persistent inbound-media storage.  Files live here so agent tools
# (vision_analyze / transcribe_audio / read_file …) can reopen them after
# ``_on_message`` returns.  Cleanup on adapter connect ages out files
# older than 24 h so the directory doesn't grow unboundedly.

def _inbound_media_dir() -> str:
    base = os.environ.get("HERMES_HOME") or os.path.expanduser("~/.hermes")
    path = os.path.join(base, "inbound_media")
    os.makedirs(path, exist_ok=True)
    return path


def _cleanup_inbound_media(older_than_seconds: int = 24 * 3600) -> int:
    """Delete inbox files older than the cutoff.  Returns the count removed."""
    removed = 0
    try:
        dir_path = _inbound_media_dir()
        now = time.time()
        for name in os.listdir(dir_path):
            full = os.path.join(dir_path, name)
            try:
                if os.path.isfile(full) and (now - os.path.getmtime(full)) > older_than_seconds:
                    os.unlink(full)
                    removed += 1
            except OSError:
                continue
    except Exception:
        logger.debug("inbound media cleanup failed", exc_info=True)
    return removed


async def _download_file_to_inbox(
    url: str, file_name: str, *, msg_id: str = "", timeout: float = 60.0,
) -> Optional[str]:
    """Download *url* into the persistent inbox; agent tools can reopen it.

    ``msg_id`` is hashed into the output filename so concurrent messages
    don't clobber each other.
    """
    if not HTTPX_AVAILABLE:
        return None
    try:
        ext = os.path.splitext(file_name)[1] or ""
        base = re.sub(r'[^\w.-]', '_', os.path.splitext(file_name)[0])[:80] or "file"
        slug = re.sub(r'[^\w]', '', msg_id)[:16] if msg_id else uuid.uuid4().hex[:12]
        out_path = os.path.join(_inbound_media_dir(), f"{slug}_{base}{ext}")
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
        with open(out_path, "wb") as fh:
            fh.write(resp.content)
        logger.info(
            "Downloaded inbound media %s (%d bytes) -> %s",
            file_name, len(resp.content), out_path,
        )
        return out_path
    except Exception as exc:
        logger.warning("Failed to download inbound media %s: %s", file_name, exc)
        return None


# Extension → MIME type used when we hand a local inbound file off to the
# agent via ``event.media_urls`` / ``media_types``.
_EXT_TO_MIME = {
    ".mp3": "audio/mpeg", ".m4a": "audio/mp4", ".wav": "audio/wav",
    ".ogg": "audio/ogg", ".amr": "audio/amr", ".aac": "audio/aac",
    ".mp4": "video/mp4", ".mov": "video/quicktime", ".webm": "video/webm",
    ".avi": "video/x-msvideo", ".mkv": "video/x-matroska",
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp", ".svg": "image/svg+xml",
    ".pdf": "application/pdf", ".md": "text/markdown", ".txt": "text/plain",
    ".csv": "text/csv", ".json": "application/json",
    ".zip": "application/zip", ".tar": "application/x-tar", ".gz": "application/gzip",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _mime_for_file(file_name: str) -> str:
    ext = os.path.splitext(file_name)[1].lower()
    return _EXT_TO_MIME.get(ext, "application/octet-stream")


async def _dingtalk_fetch_access_token(
    client_id: str, client_secret: str
) -> str:
    """Return a cached DingTalk accessToken, refreshing if near expiry.

    Uses ``/v1.0/oauth2/accessToken`` directly (no SDK dependency) so the
    ``send_message`` tool can call this without opening a Stream WebSocket
    that would conflict with the running gateway's exclusive connection.
    """
    if not HTTPX_AVAILABLE:
        raise RuntimeError("httpx not installed")

    now = datetime.now(tz=timezone.utc).timestamp()
    cached = _DINGTALK_TOKEN_CACHE.get(client_id)
    if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
        return cached[0]

    async with _DINGTALK_TOKEN_LOCK:
        cached = _DINGTALK_TOKEN_CACHE.get(client_id)
        if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
            return cached[0]
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.post(
                _DINGTALK_OAUTH_TOKEN_URL,
                json={"appKey": client_id, "appSecret": client_secret},
            )
        resp.raise_for_status()
        body = resp.json()
        token = body.get("accessToken") or ""
        expires_in = body.get("expireIn") or body.get("expires_in") or 7200
        if not token:
            raise RuntimeError(
                f"DingTalk accessToken response missing token: {body}"
            )
        _DINGTALK_TOKEN_CACHE[client_id] = (
            token, now + float(expires_in),
        )
        return token


async def _dingtalk_fetch_oapi_token(
    client_id: str, client_secret: str,
) -> Optional[str]:
    """Fetch the legacy OAPI accessToken (for /media/upload).

    DingTalk has two parallel token systems:
      - ``/v1.0/oauth2/accessToken``  — new, for message OpenAPI
      - ``/gettoken``                 — legacy OAPI, for /media/upload
    Cached separately because they're different endpoints returning
    different tokens.  Returns None on any failure (non-fatal —
    callers should fall back to text-only).
    """
    if not HTTPX_AVAILABLE:
        return None
    now = datetime.now(tz=timezone.utc).timestamp()
    cached = _DINGTALK_OAPI_TOKEN_CACHE.get(client_id)
    if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
        return cached[0]

    async with _DINGTALK_OAPI_TOKEN_LOCK:
        cached = _DINGTALK_OAPI_TOKEN_CACHE.get(client_id)
        if cached and cached[1] - _DINGTALK_TOKEN_SAFETY_MARGIN > now:
            return cached[0]
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    _DINGTALK_OAPI_TOKEN_URL,
                    params={"appkey": client_id, "appsecret": client_secret},
                )
            body = resp.json()
            if body.get("errcode") != 0 or not body.get("access_token"):
                return None
            token = str(body["access_token"])
            expires_in = float(body.get("expires_in") or 7200)
            _DINGTALK_OAPI_TOKEN_CACHE[client_id] = (
                token, now + expires_in,
            )
            return token
        except Exception:
            return None


def _classify_media_kind(path: str) -> str:
    """Return one of 'image' / 'voice' / 'video' / 'file' for a local path."""
    ext = os.path.splitext(path)[1].lower()
    if ext in _IMAGE_EXTS:
        return "image"
    if ext in _VOICE_EXTS:
        return "voice"
    if ext in _VIDEO_EXTS:
        return "video"
    return "file"


async def _dingtalk_upload_media(
    *,
    oapi_token: str,
    file_path: str,
    media_kind: str,
) -> Optional[str]:
    """Upload a local file to DingTalk /media/upload.

    Returns the raw ``media_id`` on success (preserving the leading ``@``
    when present), or None on failure.

    Caller is responsible for shaping the id per msgKey:
      - ``sampleImageMsg.photoURL``   → ``@<raw>``  (raw mediaId, NOT the
        CDN download URL — the CDN URL renders as a white placeholder in
        the DingTalk client; reference messaging.ts:714-719)
      - ``sampleFile.mediaId``        → ``@<raw>``
      - ``sampleAudio.mediaId``       → ``@<raw>``
      - ``sampleVideo.videoMediaId``  → ``@<raw>``

    Matches the reference connector's ``uploadMediaToDingTalk``
    (media/common.ts:65).  Large-file chunked upload (>20MB) not yet
    ported — callers should pre-split or downscale.
    """
    if not HTTPX_AVAILABLE:
        return None
    if not os.path.exists(file_path):
        return None
    size = os.path.getsize(file_path)
    if size > _DINGTALK_MEDIA_MAX_SIZE:
        logger.warning(
            "[dingtalk] media file %s is %d bytes, exceeds 20MB limit",
            file_path, size,
        )
        return None

    ext = os.path.splitext(file_path)[1].lower()
    if media_kind == "image":
        content_type = "image/png" if ext == ".png" else "image/jpeg"
    elif media_kind == "video":
        content_type = "video/mp4" if ext == ".mp4" else "video/quicktime"
    elif media_kind == "voice":
        content_type = "audio/mpeg" if ext == ".mp3" else "audio/amr"
    else:
        content_type = "application/octet-stream"

    # DingTalk OAPI quirk: videos must be uploaded with type=file
    # (reference: media/common.ts:141).
    upload_type = "file" if media_kind == "video" else media_kind

    try:
        with open(file_path, "rb") as fh:
            files = {
                "media": (
                    os.path.basename(file_path),
                    fh,
                    content_type,
                ),
            }
            async with httpx.AsyncClient(timeout=60.0) as client:
                resp = await client.post(
                    _DINGTALK_OAPI_MEDIA_UPLOAD_URL,
                    params={"access_token": oapi_token, "type": upload_type},
                    files=files,
                )
        body = resp.json()
        if body.get("errcode") != 0:
            logger.warning(
                "[dingtalk] /media/upload errcode=%s errmsg=%s",
                body.get("errcode"), body.get("errmsg"),
            )
            return None
        media_id = body.get("media_id") or ""
        return media_id or None
    except Exception as e:
        logger.warning("[dingtalk] /media/upload failed: %s", e)
        return None


def _dingtalk_classify_chat_id(chat_id: str) -> tuple[str, str]:
    """Classify a DingTalk chat_id into (bucket, resolved_id).

    Accepts the same target-prefix convention the dingtalk-openclaw-connector
    reference uses (``src/services/messaging.ts:sendTextToDingTalk``):

    - ``"group:<openConversationId>"`` → group
    - ``"user:<staffId>"``              → oto
    - ``"cid...=="``                    → group (auto-detect)
    - ``"$:LWCP_..."``                  → lwcp (fast-fail, needs contact perms)
    - anything else                     → oto (assume staffId)

    Returns ``(bucket, id)`` where ``id`` is the stripped value to put into
    the OpenAPI payload.
    """
    if not chat_id:
        return "oto", ""
    if chat_id.startswith("group:"):
        return "group", chat_id[6:]
    if chat_id.startswith("user:"):
        return "oto", chat_id[5:]
    if _LWCP_SENDER_RE.match(chat_id):
        return "lwcp", chat_id
    if _OPEN_CONVERSATION_RE.match(chat_id):
        return "group", chat_id
    return "oto", chat_id


# Smart markdown detection — content contains any of these characters/leading
# markers, assume markdown and upgrade msgKey to sampleMarkdown (reference:
# messaging.ts:820-826).  Plain-text prose uses sampleText so DingTalk's
# client renders it without extra padding/heading inference.
_MARKDOWN_LEADING_RE = re.compile(r'^[#*>\-]')
_MARKDOWN_INLINE_RE = re.compile(r'[*_`#\[\]]')


def _looks_like_markdown(content: str) -> bool:
    if not content:
        return False
    first_line = content.lstrip().split("\n", 1)[0]
    if _MARKDOWN_LEADING_RE.match(first_line):
        return True
    if _MARKDOWN_INLINE_RE.search(content):
        return True
    if "\n" in content:  # multi-line prose renders better as markdown
        return True
    return False


def _dingtalk_build_msg_param(
    content: str,
    *,
    title: str = "Hermes",
    msg_type: Optional[str] = None,
) -> tuple[str, str]:
    """Return (msgKey, JSON-encoded msgParam).

    ``msg_type`` overrides auto-detection:
      - "text"      → sampleText
      - "markdown"  → sampleMarkdown
      - "image"     → sampleImageMsg (content = mediaId or photoURL)
      - "file"      → sampleFile (content = mediaId)
      - "voice"     → sampleAudio (content = mediaId; duration=0 per reference)
      - "video"     → sampleVideo (content = mediaId)
    If None, chooses sampleMarkdown vs sampleText via
    ``_looks_like_markdown``.  Matches the reference's ``buildMsgPayload``
    (messaging.ts:150+).
    """
    truncated = content[:MAX_MESSAGE_LENGTH]

    if msg_type == "image":
        # sampleImageMsg.photoURL accepts the raw media_id with the leading
        # ``@`` preserved — DingTalk's server-side resolves this to the
        # tenant-internal image store.  Reference (messaging.ts:714-719 +
        # messaging.ts:188-192): `photoURL: uploadResult.mediaId` with the
        # inline comment "使用原始 mediaId（带 @）".
        #
        # Earlier iterations passed `https://down.dingtalk.com/media/<clean>`
        # here.  The API accepted it, but DingTalk clients rendered a blank
        # (white) image because the CDN URL requires a tenant-scoped auth
        # handshake that the client doesn't perform during `sampleImageMsg`
        # rendering.  Only keep the URL passthrough for already-public http(s)
        # URLs the caller provides directly.
        if truncated.startswith(("http://", "https://")):
            photo_url = truncated
        else:
            photo_url = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleImageMsg", json.dumps(
            {"photoURL": photo_url}, ensure_ascii=False,
        )
    if msg_type == "file":
        # sampleFile / sampleAudio / sampleVideo take the raw media_id with
        # leading ``@`` preserved (reference: media.ts:680+ & :876).
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleFile", json.dumps(
            {"mediaId": media_id, "fileName": title, "fileType": "file"},
            ensure_ascii=False,
        )
    if msg_type == "voice":
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleAudio", json.dumps(
            {"mediaId": media_id, "duration": "0"}, ensure_ascii=False,
        )
    if msg_type == "video":
        media_id = truncated if truncated.startswith("@") else f"@{truncated}"
        return "sampleVideo", json.dumps(
            {"videoMediaId": media_id, "videoType": "mp4",
             "picMediaId": "", "duration": "0"},
            ensure_ascii=False,
        )
    if msg_type == "text":
        return "sampleText", json.dumps(
            {"content": truncated}, ensure_ascii=False,
        )
    if msg_type == "markdown" or _looks_like_markdown(truncated):
        return "sampleMarkdown", json.dumps(
            {"title": title, "text": truncated}, ensure_ascii=False,
        )
    return "sampleText", json.dumps(
        {"content": truncated}, ensure_ascii=False,
    )


async def _dingtalk_post_one(
    *,
    token: str,
    robot_code: str,
    bucket: str,
    target_id: str,
    msg_key: str,
    msg_param: str,
) -> Dict[str, Any]:
    """POST a single already-built message to the correct Robot OpenAPI endpoint."""
    if bucket == "group":
        url = _DINGTALK_GROUP_SEND_URL
        payload = {
            "robotCode": robot_code,
            "openConversationId": target_id,
            "msgKey": msg_key,
            "msgParam": msg_param,
        }
    else:
        url = _DINGTALK_OTO_BATCH_SEND_URL
        payload = {
            "robotCode": robot_code,
            "userIds": [target_id],
            "msgKey": msg_key,
            "msgParam": msg_param,
        }

    headers = {
        "x-acs-dingtalk-access-token": token,
        "Content-Type": "application/json",
    }
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(url, json=payload, headers=headers)
    except Exception as e:
        return {"error": f"DingTalk OpenAPI request failed: {e}"}

    if resp.status_code >= 300:
        return {
            "error": f"DingTalk OpenAPI HTTP {resp.status_code}: {resp.text[:300]}"
        }

    try:
        body = resp.json()
    except Exception:
        body = {}
    # Robot OpenAPI returns ``processQueryKey`` on success (reference:
    # openclaw-connector messaging.ts:254,331,999).  Fallback to requestId.
    request_id = body.get("processQueryKey") or body.get("requestId") or ""
    return {"success": True, "request_id": request_id}


async def dingtalk_send_proactive(
    *,
    client_id: str,
    client_secret: str,
    robot_code: str,
    chat_id: str,
    content: str,
    media_files: Optional[List[tuple[str, bool]]] = None,
    title: str = "Hermes",
) -> Dict[str, Any]:
    """Send a message via DingTalk Robot OpenAPI without Stream Mode.

    Proactive-send counterpart to ``DingTalkAdapter.send()``'s
    session_webhook path.  Authenticates with AppKey/Secret, routes by
    ``chat_id`` shape (group / oto / lwcp), and optionally uploads+sends
    media attachments as additional messages.

    ``media_files`` is a list of ``(local_path, is_voice_hint)`` tuples.
    Each file gets uploaded to /media/upload, then sent as a dedicated
    message (sampleImageMsg / sampleAudio / sampleVideo / sampleFile).
    Text is sent first (if non-empty), then each media in order.  Mirrors
    Feishu's approach of emitting separate messages per attachment.

    ``title`` is used for sampleMarkdown headings when content auto-upgrades.

    Returns ``{"success": True, "request_id": ..., "route": ...}`` on success
    of the primary (text or first) message.  ``media_results`` lists each
    media attempt with its outcome.
    """
    if not HTTPX_AVAILABLE:
        return {"error": "httpx not installed"}
    if not client_id or not client_secret:
        return {"error": "DingTalk proactive send requires client_id + client_secret"}
    if not robot_code:
        return {"error": "DingTalk proactive send requires robot_code (defaults to client_id)"}
    if not chat_id:
        return {"error": "DingTalk proactive send requires chat_id"}

    # Classify FIRST so LWCP fast-fails with zero network roundtrip.
    bucket, target_id = _dingtalk_classify_chat_id(chat_id)
    if bucket == "lwcp":
        return {
            "error": (
                "DingTalk chat_id is LWCP-encoded sender_id (%s…): cannot be routed "
                "directly by Robot OpenAPI. Grant the app 'qyapi_get_department_list' "
                "+ 'qyapi_get_department_member' permissions at open-dev.dingtalk.com "
                "and resolve to plain staffId before calling proactive send."
            ) % target_id[:16]
        }

    try:
        token = await _dingtalk_fetch_access_token(client_id, client_secret)
    except Exception as e:
        return {"error": f"DingTalk accessToken fetch failed: {e}"}

    results: List[Dict[str, Any]] = []
    primary_request_id = ""

    # 1. Text first (if non-empty).  Smart markdown detection applied here.
    if content and content.strip():
        msg_key, msg_param = _dingtalk_build_msg_param(content, title=title)
        text_result = await _dingtalk_post_one(
            token=token, robot_code=robot_code, bucket=bucket,
            target_id=target_id, msg_key=msg_key, msg_param=msg_param,
        )
        if text_result.get("error"):
            return text_result
        primary_request_id = text_result.get("request_id", "")

    # 2. Media (one message per file).  Non-fatal per file — we collect
    # per-file results and let the caller decide what to surface.
    if media_files:
        oapi_token = await _dingtalk_fetch_oapi_token(client_id, client_secret)
        if not oapi_token:
            results.append({
                "error": (
                    "Could not obtain OAPI token for /media/upload — verify the "
                    "app has ServerAPI permissions (qyapi_media) and retry."
                ),
            })
            # Keep the text result if we had one.
            if primary_request_id:
                return {
                    "success": True,
                    "request_id": primary_request_id,
                    "chat_id": target_id,
                    "route": bucket,
                    "media_results": results,
                }
            return results[-1]

        for raw_path, is_voice_hint in media_files:
            media_kind = (
                "voice" if is_voice_hint and _classify_media_kind(raw_path) in ("voice", "file")
                else _classify_media_kind(raw_path)
            )
            media_id = await _dingtalk_upload_media(
                oapi_token=oapi_token,
                file_path=raw_path,
                media_kind=media_kind,
            )
            if not media_id:
                results.append({
                    "error": f"upload failed for {os.path.basename(raw_path)}",
                    "path": raw_path,
                })
                continue
            m_msg_key, m_msg_param = _dingtalk_build_msg_param(
                media_id,
                msg_type=media_kind,
                title=os.path.basename(raw_path),
            )
            m_result = await _dingtalk_post_one(
                token=token, robot_code=robot_code, bucket=bucket,
                target_id=target_id, msg_key=m_msg_key, msg_param=m_msg_param,
            )
            m_result["kind"] = media_kind
            m_result["path"] = raw_path
            results.append(m_result)
            if not primary_request_id and m_result.get("success"):
                primary_request_id = m_result.get("request_id", "")

    if not primary_request_id and not content.strip():
        # Nothing delivered (empty text, all media failed).
        return {
            "error": "Nothing delivered: empty text and all media uploads failed",
            "media_results": results,
        }

    return {
        "success": True,
        "request_id": primary_request_id,
        "chat_id": target_id,
        "route": bucket,
        **({"media_results": results} if results else {}),
    }


def check_dingtalk_requirements() -> bool:
    """Check if DingTalk dependencies are available and configured.

    Lazy-installs dingtalk-stream via ``tools.lazy_deps.ensure("platform.dingtalk")``
    on first call if not present.
    """
    global DINGTALK_STREAM_AVAILABLE, dingtalk_stream, ChatbotMessage, CallbackMessage, AckMessage
    global HTTPX_AVAILABLE, httpx
    if not DINGTALK_STREAM_AVAILABLE or not HTTPX_AVAILABLE:
        try:
            from tools.lazy_deps import ensure as _lazy_ensure
            _lazy_ensure("platform.dingtalk", prompt=False)
        except Exception:
            return False
        try:
            import dingtalk_stream as _ds
            from dingtalk_stream import ChatbotMessage as _CM
            from dingtalk_stream.frames import CallbackMessage as _CBM, AckMessage as _AM
            import httpx as _httpx
        except Exception:
            return False
        dingtalk_stream = _ds
        ChatbotMessage = _CM
        CallbackMessage = _CBM
        AckMessage = _AM
        httpx = _httpx
        DINGTALK_STREAM_AVAILABLE = True
        HTTPX_AVAILABLE = True
    if not os.getenv("DINGTALK_CLIENT_ID") or not os.getenv("DINGTALK_CLIENT_SECRET"):
        return False
    return True


class DingTalkAdapter(BasePlatformAdapter):
    """DingTalk chatbot adapter using Stream Mode.

    The dingtalk-stream SDK maintains a long-lived WebSocket connection.
    Incoming messages arrive via a ChatbotHandler callback. Replies are
    sent via the incoming message's session_webhook URL using httpx.

    Features:
    - Text messages (plain + rich text)
    - Images, audio, video, files (via download codes)
    - Group chat @mention detection
    - Session webhook caching with expiry tracking
    - Markdown formatted replies
    """

    MAX_MESSAGE_LENGTH = MAX_MESSAGE_LENGTH

    @property
    def SUPPORTS_MESSAGE_EDITING(self) -> bool:  # noqa: N802
        """Edits only meaningful when AI Cards are configured.

        The gateway gates streaming cursor + edit behaviour on this flag,
        so we must reflect the actual adapter capability at runtime.
        """
        return bool(self._card_template_id and self._card_sdk)

    @property
    def REQUIRES_EDIT_FINALIZE(self) -> bool:  # noqa: N802
        """AI Card lifecycle requires an explicit ``finalize=True`` edit
        to close the streaming indicator, even when the final content is
        identical to the last streamed update.  Enabled only when cards
        are configured — webhook-only DingTalk doesn't need it.
        """
        return bool(self._card_template_id and self._card_sdk)

    def __init__(self, config: PlatformConfig):
        super().__init__(config, Platform.DINGTALK)

        extra = config.extra or {}
        self._client_id: str = extra.get("client_id") or os.getenv(
            "DINGTALK_CLIENT_ID", ""
        )
        self._client_secret: str = extra.get("client_secret") or os.getenv(
            "DINGTALK_CLIENT_SECRET", ""
        )

        # Group-chat gating (mirrors Slack/Telegram/Discord/WhatsApp conventions).
        # Mention state is the structured ``is_in_at_list`` attribute from the
        # dingtalk-stream SDK (set from the callback's ``isInAtList`` flag),
        # not text parsing.
        self._mention_patterns: List[re.Pattern] = self._compile_mention_patterns()
        self._allowed_users: Set[str] = self._load_allowed_users()

        self._stream_client: Any = None
        self._stream_task: Optional[asyncio.Task] = None
        self._http_client: Optional["httpx.AsyncClient"] = None
        self._card_sdk: Optional[Any] = None
        self._robot_sdk: Optional[Any] = None
        self._robot_code: str = extra.get("robot_code") or self._client_id

        # Message deduplication
        self._dedup = MessageDeduplicator(max_size=1000)
        # Map chat_id -> (session_webhook, expired_time_ms) for reply routing
        self._session_webhooks: Dict[str, tuple[str, int]] = {}
        # Map chat_id -> last inbound ChatbotMessage. Keyed by chat_id instead
        # of a single class attribute to avoid cross-message clobbering when
        # multiple conversations run concurrently.
        self._message_contexts: Dict[str, Any] = {}
        self._card_template_id: Optional[str] = extra.get("card_template_id")

        # Chats for which we've already fired the Done reaction — prevents
        # double-firing across segment boundaries or parallel flows
        # (tool-progress + stream-consumer both finalizing their cards).
        # Reset each inbound message.
        self._done_emoji_fired: Set[str] = set()
        # Cards in streaming state per chat: chat_id -> { out_track_id -> last_content }.
        # Every `send()` creates+finalizes a card (closed state).  A subsequent
        # `edit_message(finalize=False)` re-opens the card (DingTalk's API
        # allows streaming_update on a finalized card — it flips back to
        # streaming).  We track those reopened cards so the next `send()` can
        # auto-close them as siblings — otherwise tool-progress cards get
        # stuck in streaming state forever.
        self._streaming_cards: Dict[str, Dict[str, str]] = {}
        # Per-card last-edit timestamp (ms) — used to throttle non-finalize
        # edit_message() calls so DingTalk's concurrent-update 403 doesn't
        # fire.  Finalize edits bypass the check.  Cleared on finalize.
        self._card_last_edit_ms: Dict[str, int] = {}
        # Per-chat error-send cooldown (chat_id -> last error send ts ms)
        # so a burst of failures doesn't spam the user.
        self._error_last_sent_ms: Dict[str, int] = {}
        # Track fire-and-forget emoji/reaction coroutines so Python's GC
        # doesn't drop them mid-flight, and we can cancel them on disconnect.
        self._bg_tasks: Set[asyncio.Task] = set()

        # Per-session inbound message queue (promise-chain pattern ported
        # from the reference connector, core/message-handler.ts:86-105).
        # queueKey is the chat_id (conversation_id or sender_id).  Each
        # new inbound message chains a ``.then()`` onto the previous
        # task so same-session messages are processed in arrival order,
        # avoiding parallel agent runs that race on the same context.
        self._session_queues: Dict[str, asyncio.Task] = {}
        self._session_last_activity: Dict[str, float] = {}
        # Periodic sweep of stale queue entries (>5 min idle) so long-
        # idle chats don't keep references to old Tasks forever.
        self._session_queue_sweeper: Optional[asyncio.Task] = None

    # -- Connection lifecycle -----------------------------------------------

    async def connect(self) -> bool:
        """Connect to DingTalk via Stream Mode."""
        # Age out stale inbound-media downloads so the directory doesn't grow
        # unboundedly.  Safe every connect — O(files), threshold (24 h) is
        # well above any reasonable agent session.
        removed = _cleanup_inbound_media()
        if removed:
            logger.info(
                "[%s] Cleaned up %d stale inbound-media file(s)",
                self.name, removed,
            )

        if not DINGTALK_STREAM_AVAILABLE:
            logger.warning(
                "[%s] dingtalk-stream not installed. Run: pip install 'dingtalk-stream>=0.20'",
                self.name,
            )
            return False
        if not HTTPX_AVAILABLE:
            logger.warning(
                "[%s] httpx not installed. Run: pip install httpx", self.name
            )
            return False
        if not self._client_id or not self._client_secret:
            logger.warning(
                "[%s] DINGTALK_CLIENT_ID and DINGTALK_CLIENT_SECRET required", self.name
            )
            return False

        try:
            # Tighter keepalive so idle CLOSE_WAIT drains promptly (#18451).
            from gateway.platforms._http_client_limits import platform_httpx_limits
            self._http_client = httpx.AsyncClient(
                timeout=30.0, limits=platform_httpx_limits(),
            )

            _install_dingtalk_websockets_proxy()

            credential = dingtalk_stream.Credential(
                self._client_id, self._client_secret
            )
            self._stream_client = dingtalk_stream.DingTalkStreamClient(credential)

            # Initialize card SDK if available and configured
            if CARD_SDK_AVAILABLE and self._card_template_id:
                sdk_config = open_api_models.Config()
                sdk_config.protocol = "https"
                sdk_config.region_id = "central"
                self._card_sdk = dingtalk_card_client.Client(sdk_config)
                self._robot_sdk = dingtalk_robot_client.Client(sdk_config)
                logger.info(
                    "[%s] Card SDK initialized with template: %s",
                    self.name,
                    self._card_template_id,
                )
            elif CARD_SDK_AVAILABLE:
                # Initialize robot SDK even without card template (for media download)
                sdk_config = open_api_models.Config()
                sdk_config.protocol = "https"
                sdk_config.region_id = "central"
                self._robot_sdk = dingtalk_robot_client.Client(sdk_config)
                logger.info("[%s] Robot SDK initialized (media download)", self.name)

            # Capture the current event loop for cross-thread dispatch
            loop = asyncio.get_running_loop()
            handler = _IncomingHandler(self, loop)
            self._stream_client.register_callback_handler(
                dingtalk_stream.ChatbotMessage.TOPIC, handler
            )

            self._stream_task = asyncio.create_task(self._run_stream())
            # Start the periodic sweeper for stale inbound-queue entries.
            # Mirrors core/message-handler.ts:105 (setInterval 60s).
            self._session_queue_sweeper = asyncio.create_task(
                self._sweep_session_queues()
            )
            self._mark_connected()
            logger.info("[%s] Connected via Stream Mode", self.name)
            return True
        except Exception as e:
            logger.error("[%s] Failed to connect: %s", self.name, e)
            return False

    async def _run_stream(self) -> None:
        """Run the async stream client with auto-reconnection."""
        backoff_idx = 0
        while self._running:
            try:
                logger.debug("[%s] Starting stream client...", self.name)
                await self._stream_client.start()
            except asyncio.CancelledError:
                return
            except Exception as e:
                if not self._running:
                    return
                logger.warning("[%s] Stream client error: %s", self.name, e)

            if not self._running:
                return

            delay = RECONNECT_BACKOFF[min(backoff_idx, len(RECONNECT_BACKOFF) - 1)]
            logger.info("[%s] Reconnecting in %ds...", self.name, delay)
            await asyncio.sleep(delay)
            backoff_idx += 1

    async def disconnect(self) -> None:
        """Disconnect from DingTalk."""
        self._running = False
        self._mark_disconnected()

        # Close the active websocket first so the stream task sees the
        # disconnection and exits cleanly, rather than getting stuck
        # awaiting frames that will never arrive.
        websocket = getattr(self._stream_client, "websocket", None) if self._stream_client else None
        if websocket is not None:
            try:
                await websocket.close()
            except Exception as e:
                logger.debug("[%s] websocket close during disconnect failed: %s", self.name, e)

        if self._stream_task:
            # Try graceful close first if SDK supports it. The SDK's close()
            # is sync and may block on network I/O, so offload to a thread.
            if hasattr(self._stream_client, "close"):
                try:
                    await asyncio.to_thread(self._stream_client.close)
                except Exception:
                    pass

            self._stream_task.cancel()
            try:
                await asyncio.wait_for(self._stream_task, timeout=5.0)
            except (asyncio.CancelledError, asyncio.TimeoutError):
                logger.debug("[%s] stream task did not exit cleanly during disconnect", self.name)
            self._stream_task = None

        # Stop the session-queue sweeper.
        if self._session_queue_sweeper:
            self._session_queue_sweeper.cancel()
            try:
                await self._session_queue_sweeper
            except (asyncio.CancelledError, Exception):
                pass
            self._session_queue_sweeper = None

        # Cancel any still-pending inbound-queue tasks so disconnect
        # doesn't hang waiting on half-handled messages.
        if self._session_queues:
            for task in list(self._session_queues.values()):
                if not task.done():
                    task.cancel()
            await asyncio.gather(
                *self._session_queues.values(), return_exceptions=True,
            )
            self._session_queues.clear()
        self._session_last_activity.clear()

        # Cancel any in-flight background tasks (emoji reactions, etc.)
        if self._bg_tasks:
            for task in list(self._bg_tasks):
                task.cancel()
            await asyncio.gather(*self._bg_tasks, return_exceptions=True)
            self._bg_tasks.clear()

        # Finalize any open streaming cards before the HTTP client closes so
        # they don't stay stuck in streaming state on DingTalk's UI after
        # a gateway restart.  _close_streaming_siblings handles its own
        # per-card exceptions; the outer try is a safety net for token fetch.
        for _chat_id in list(self._streaming_cards):
            try:
                await self._close_streaming_siblings(_chat_id)
            except Exception as _exc:
                logger.debug(
                    "[%s] Failed to finalize streaming card on disconnect for %s: %s",
                    self.name, _chat_id, _exc,
                )

        if self._http_client:
            await self._http_client.aclose()
            self._http_client = None

        self._stream_client = None
        self._session_webhooks.clear()
        self._message_contexts.clear()
        self._streaming_cards.clear()
        self._card_last_edit_ms.clear()
        self._error_last_sent_ms.clear()
        self._done_emoji_fired.clear()
        self._dedup.clear()
        logger.info("[%s] Disconnected", self.name)

    # -- Group gating --------------------------------------------------------

    def _dingtalk_require_mention(self) -> bool:
        """Return whether group chats should require an explicit bot trigger."""
        configured = self.config.extra.get("require_mention")
        if configured is not None:
            if isinstance(configured, str):
                return configured.lower() in {"true", "1", "yes", "on"}
            return bool(configured)
        return os.getenv("DINGTALK_REQUIRE_MENTION", "false").lower() in {"true", "1", "yes", "on"}

    def _dingtalk_free_response_chats(self) -> Set[str]:
        raw = self.config.extra.get("free_response_chats")
        if raw is None:
            raw = os.getenv("DINGTALK_FREE_RESPONSE_CHATS", "")
        if isinstance(raw, list):
            return {str(part).strip() for part in raw if str(part).strip()}
        return {part.strip() for part in str(raw).split(",") if part.strip()}

    def _dingtalk_allowed_chats(self) -> Set[str]:
        """Return the whitelist of group chat IDs the bot will respond in.

        When non-empty, group messages from chats NOT in this set are silently
        ignored — even if the bot is @mentioned.  DMs are never filtered.
        Empty set means no restriction (fully backward compatible).
        """
        raw = self.config.extra.get("allowed_chats") if self.config.extra else None
        if raw is None:
            raw = os.getenv("DINGTALK_ALLOWED_CHATS", "")
        if isinstance(raw, list):
            return {str(part).strip() for part in raw if str(part).strip()}
        return {part.strip() for part in str(raw).split(",") if part.strip()}

    def _compile_mention_patterns(self) -> List[re.Pattern]:
        """Compile optional regex wake-word patterns for group triggers."""
        patterns = self.config.extra.get("mention_patterns") if self.config.extra else None
        if patterns is None:
            raw = os.getenv("DINGTALK_MENTION_PATTERNS", "").strip()
            if raw:
                try:
                    loaded = json.loads(raw)
                except Exception:
                    loaded = [part.strip() for part in raw.splitlines() if part.strip()]
                    if not loaded:
                        loaded = [part.strip() for part in raw.split(",") if part.strip()]
                patterns = loaded

        if patterns is None:
            return []
        if isinstance(patterns, str):
            patterns = [patterns]
        if not isinstance(patterns, list):
            logger.warning(
                "[%s] dingtalk mention_patterns must be a list or string; got %s",
                self.name,
                type(patterns).__name__,
            )
            return []

        compiled: List[re.Pattern] = []
        for pattern in patterns:
            if not isinstance(pattern, str) or not pattern.strip():
                continue
            try:
                compiled.append(re.compile(pattern, re.IGNORECASE))
            except re.error as exc:
                logger.warning("[%s] Invalid DingTalk mention pattern %r: %s", self.name, pattern, exc)
        if compiled:
            logger.info("[%s] Loaded %d DingTalk mention pattern(s)", self.name, len(compiled))
        return compiled

    def _load_allowed_users(self) -> Set[str]:
        """Load allowed-users list from config.extra or env var.

        IDs are matched case-insensitively against the sender's ``staff_id`` and
        ``sender_id``. A wildcard ``*`` disables the check.
        """
        raw = self.config.extra.get("allowed_users") if self.config.extra else None
        if raw is None:
            raw = os.getenv("DINGTALK_ALLOWED_USERS", "")
        if isinstance(raw, list):
            items = [str(part).strip() for part in raw if str(part).strip()]
        else:
            items = [part.strip() for part in str(raw).split(",") if part.strip()]
        return {item.lower() for item in items}

    def _is_user_allowed(self, sender_id: str, sender_staff_id: str) -> bool:
        if not self._allowed_users or "*" in self._allowed_users:
            return True
        candidates = {(sender_id or "").lower(), (sender_staff_id or "").lower()}
        candidates.discard("")
        return bool(candidates & self._allowed_users)

    def _message_mentions_bot(self, message: "ChatbotMessage") -> bool:
        """True if the bot was @-mentioned in a group message.

        dingtalk-stream sets ``is_in_at_list`` on the incoming ChatbotMessage
        when the bot is addressed via @-mention.
        """
        return bool(getattr(message, "is_in_at_list", False))

    def _message_matches_mention_patterns(self, text: str) -> bool:
        if not text or not self._mention_patterns:
            return False
        return any(pattern.search(text) for pattern in self._mention_patterns)

    def _should_process_message(self, message: "ChatbotMessage", text: str, is_group: bool, chat_id: str) -> bool:
        """Apply DingTalk group trigger rules.

        DMs remain unrestricted (subject to ``allowed_users`` which is enforced
        earlier). Group messages are accepted when:
        - the chat passes the ``allowed_chats`` whitelist (when set)
        - the chat is explicitly allowlisted in ``free_response_chats``
        - ``require_mention`` is disabled
        - the bot is @mentioned (``is_in_at_list``)
        - the text matches a configured regex wake-word pattern

        When ``allowed_chats`` is non-empty, it acts as a hard gate — messages
        from any group chat not in the list are ignored regardless of the
        other rules.
        """
        if not is_group:
            return True
        allowed = self._dingtalk_allowed_chats()
        if allowed and chat_id and chat_id not in allowed:
            return False
        if chat_id and chat_id in self._dingtalk_free_response_chats():
            return True
        if not self._dingtalk_require_mention():
            return True
        if self._message_mentions_bot(message):
            return True
        return self._message_matches_mention_patterns(text)

    def _spawn_bg(self, coro) -> None:
        """Start a fire-and-forget coroutine and track it for cleanup."""
        task = asyncio.create_task(coro)
        self._bg_tasks.add(task)
        task.add_done_callback(self._bg_tasks.discard)

    # -- Inbound serialization queue ---------------------------------------

    def _inbound_queue_key(self, chatbot_msg: "ChatbotMessage") -> str:
        """Build the queueKey for same-session inbound serialization.

        Mirrors core/message-handler.ts:1601-1649 but simplified: hermes
        has one agent per adapter, so we key on chat_id alone
        (conversation_id for groups, sender_id for DMs).  Parallel
        different-chat messages still fan out — the queue only serializes
        same-chat arrivals.
        """
        conv_id = getattr(chatbot_msg, "conversation_id", "") or ""
        sender_id = getattr(chatbot_msg, "sender_id", "") or ""
        return conv_id or sender_id

    async def _send_busy_ack(self, chatbot_msg: "ChatbotMessage") -> None:
        """Send a text ACK when a new message lands on a busy queue.

        Reference: core/message-handler.ts:1662-1689 creates an AI Card
        with ACK text and hands it to the dispatcher for later reuse.
        hermes does NOT thread a pre-created card into ``send()`` (that
        refactor is out of scope for this port), so we fall back to a
        plain text webhook reply — users still see immediate feedback
        that the bot received their message, just as a separate bubble
        instead of an in-place card update.
        """
        if not self._http_client:
            return
        webhook = getattr(chatbot_msg, "session_webhook", "") or ""
        if not webhook or not _DINGTALK_WEBHOOK_RE.match(webhook):
            return
        phrase = random.choice(_QUEUE_BUSY_ACK_PHRASES)
        try:
            await self._http_client.post(
                webhook,
                json={"msgtype": "text", "text": {"content": phrase}},
                timeout=5.0,
            )
        except Exception as e:
            logger.debug("[%s] busy ACK send failed: %s", self.name, e)

    async def _enqueue_inbound(self, chatbot_msg: "ChatbotMessage") -> None:
        """Chain inbound message processing onto the per-chat queue tail.

        Mirrors core/message-handler.ts:1651-1719.  Same-chat messages are
        processed strictly in order; different chats run in parallel.
        When a new message lands while the queue is already working on a
        previous one, a busy-ACK is sent fire-and-forget so the user sees
        immediate feedback that we noticed their message.
        """
        queue_key = self._inbound_queue_key(chatbot_msg)
        if not queue_key:
            await self._on_message(chatbot_msg)
            return

        self._session_last_activity[queue_key] = time.monotonic()
        prev_task = self._session_queues.get(queue_key)
        is_busy = prev_task is not None and not prev_task.done()

        if is_busy:
            self._spawn_bg(self._send_busy_ack(chatbot_msg))

        async def _chained() -> None:
            if prev_task is not None:
                try:
                    await prev_task
                except Exception:
                    # Prior task's failure is its own problem; don't
                    # block the whole queue.
                    pass
            await self._on_message(chatbot_msg)

        task = asyncio.create_task(_chained())
        self._session_queues[queue_key] = task

        def _cleanup(t: asyncio.Task) -> None:
            # Only clear the map if we're still the tail — a newer
            # enqueue may have already replaced us.
            if self._session_queues.get(queue_key) is t:
                self._session_queues.pop(queue_key, None)

        task.add_done_callback(_cleanup)

    async def _sweep_session_queues(self) -> None:
        """Periodic sweep of stale queueKey entries.

        Runs every 60s.  Entries idle for more than ``_INBOUND_QUEUE_TTL_SEC``
        are dropped from the activity map; entries whose Task already
        completed are also dropped from ``_session_queues`` (defensive —
        done_callback normally handles this, but an entry may survive if
        the callback raced with sweep).  Mirrors
        core/message-handler.ts:94-105.
        """
        try:
            while self._running:
                await asyncio.sleep(60)
                now = time.monotonic()
                stale = [
                    k for k, ts in self._session_last_activity.items()
                    if now - ts > _INBOUND_QUEUE_TTL_SEC
                ]
                for k in stale:
                    self._session_last_activity.pop(k, None)
                    task = self._session_queues.get(k)
                    if task is not None and task.done():
                        self._session_queues.pop(k, None)
        except asyncio.CancelledError:
            return
        except Exception as e:
            logger.debug("[%s] session-queue sweep error: %s", self.name, e)

    # -- AI Card lifecycle helpers ------------------------------------------

    async def _close_streaming_siblings(self, chat_id: str) -> None:
        """Finalize any previously-open streaming cards for this chat.

        Called at the start of every ``send()`` so lingering tool-progress
        cards that were reopened by ``edit_message(finalize=False)`` get
        cleanly closed before the next card is created.  Without this,
        tool-progress cards stay stuck in streaming state after the agent
        moves on (there is no explicit "turn end" signal from the gateway).
        """
        cards = self._streaming_cards.pop(chat_id, None)
        if not cards:
            return
        token = await self._get_access_token()
        if not token:
            return
        for out_track_id, last_content in list(cards.items()):
            try:
                await self._stream_card_content(
                    out_track_id, token, last_content, finalize=True,
                )
                logger.debug(
                    "[%s] AI Card sibling closed: %s",
                    self.name, out_track_id,
                )
            except Exception as e:
                logger.debug(
                    "[%s] Sibling close failed for %s: %s",
                    self.name, out_track_id, e,
                )

    def _fire_done_reaction(self, chat_id: str) -> None:
        """Swap 🤔Thinking → 🥳Done on the original user message.

        Idempotent per chat_id — safe to call from segment-break flushes
        and final-done flushes without double-firing.
        """
        if chat_id in self._done_emoji_fired:
            return
        self._done_emoji_fired.add(chat_id)
        msg = self._message_contexts.get(chat_id)
        if not msg:
            return
        msg_id = getattr(msg, "message_id", "") or ""
        conversation_id = getattr(msg, "conversation_id", "") or ""
        if not (msg_id and conversation_id):
            return

        async def _swap() -> None:
            await self._send_emotion(
                msg_id, conversation_id, "🤔Thinking", recall=True,
            )
            await self._send_emotion(
                msg_id, conversation_id, "🥳Done", recall=False,
            )

        self._spawn_bg(_swap())

    # -- Inbound message processing -----------------------------------------

    async def _on_message(
        self,
        message: "ChatbotMessage",
    ) -> None:
        """Process an incoming DingTalk chatbot message."""
        msg_id = getattr(message, "message_id", None) or uuid.uuid4().hex
        if self._dedup.is_duplicate(msg_id):
            logger.debug("[%s] Duplicate message %s, skipping", self.name, msg_id)
            return

        # Chat context
        conversation_id = getattr(message, "conversation_id", "") or ""
        conversation_type = getattr(message, "conversation_type", "1")
        is_group = str(conversation_type) == "2"
        sender_id = getattr(message, "sender_id", "") or ""
        sender_nick = getattr(message, "sender_nick", "") or sender_id
        sender_staff_id = getattr(message, "sender_staff_id", "") or ""

        chat_id = conversation_id or sender_id
        chat_type = "group" if is_group else "dm"

        # Allowed-users gate (applies to both DM and group)
        if not self._is_user_allowed(sender_id, sender_staff_id):
            logger.debug(
                "[%s] Dropping message from non-allowlisted user staff_id=%s sender_id=%s",
                self.name, sender_staff_id, sender_id,
            )
            return

        # Group mention/pattern gate.  DMs pass through unconditionally.
        # We need the message text for regex wake-word matching; extract it
        # early but don't consume the rest of the pipeline until after the
        # gate decides whether to process.
        _early_text = self._extract_text(message) or ""
        if not self._should_process_message(message, _early_text, is_group, chat_id):
            logger.debug(
                "[%s] Dropping group message that failed mention gate message_id=%s chat_id=%s",
                self.name, msg_id, chat_id,
            )
            return

        # Stash the incoming message keyed by chat_id so concurrent
        # conversations don't clobber each other's context.  Also reset
        # the per-chat "Done emoji fired" marker so a new inbound message
        # gets its own Thinking→Done cycle.
        if chat_id:
            self._message_contexts[chat_id] = message
            self._done_emoji_fired.discard(chat_id)

        # Store session webhook
        session_webhook = getattr(message, "session_webhook", None) or ""
        session_webhook_expired_time = (
            getattr(message, "session_webhook_expired_time", 0) or 0
        )
        if session_webhook and chat_id and _DINGTALK_WEBHOOK_RE.match(session_webhook):
            if len(self._session_webhooks) >= _SESSION_WEBHOOKS_MAX:
                try:
                    self._session_webhooks.pop(next(iter(self._session_webhooks)))
                except StopIteration:
                    pass
            self._session_webhooks[chat_id] = (
                session_webhook,
                session_webhook_expired_time,
            )

        # Resolve media download codes to URLs so vision tools can use them
        await self._resolve_media_codes(message)

        # Extract text content
        text = self._extract_text(message)

        # Determine message type and build media list
        msg_type, media_urls, media_types = self._extract_media(message)

        logger.info(
            "[%s] Pre-file-parse state: msg_type_str=%s text=%r media_urls=%d extensions_keys=%s",
            self.name,
            getattr(message, "message_type", "?"),
            (text[:60] if text else ""),
            len(media_urls),
            list(getattr(message, "extensions", {}).keys())[:10],
        )

        # ------------------------------------------------------------------
        # Download remote image URLs to local files.
        #
        # DingTalk's signed OSS URLs are not universally accessible — the
        # vision_analyze_tool fails when it tries to download them directly
        # because the OSS signature is bound to specific request context.
        # Mirrors connector's downloadImageToFile / downloadMediaByCode
        # approach: download in the gateway process (same IP/context as the
        # DingTalk SDK) and pass local file paths to the agent.
        # ------------------------------------------------------------------
        if media_urls:
            try:
                media_urls = await self._download_images_to_local(
                    media_urls, media_types, message,
                )
            except Exception:
                logger.warning(
                    "[%s] Image download to local failed (non-fatal), keeping URLs",
                    self.name, exc_info=True,
                )

        # ------------------------------------------------------------------
        # File content auto-parsing: download text-type files (.md, .txt,
        # .json, .docx, .pdf, etc.) and inject their content into ``text``
        # so the LLM can see the content immediately without needing tools.
        # Mirrors dingtalk-openclaw-connector core/message-handler.ts:1226-1328.
        # ------------------------------------------------------------------
        try:
            file_parts, file_attachments = await self._extract_and_parse_file_attachments(message)
            if file_parts:
                file_text = "\n\n".join(file_parts)
                text = f"{text}\n\n{file_text}" if text else file_text
                logger.info(
                    "[%s] Injected %d file content block(s) into message text",
                    self.name, len(file_parts),
                )
            # Expose each downloaded media file to the agent as a ``file://``
            # URI so tools (transcribe_audio / vision_analyze / read_file …)
            # can open them.  ``_extract_media`` + ``_download_images_to_local``
            # already populated ``media_urls`` for image/richText items; this
            # appends the audio/video/file items we just downloaded.
            for local_path, mime in file_attachments:
                media_urls.append(f"file://{local_path}")
                media_types.append(mime)
            if file_attachments:
                logger.info(
                    "[%s] Attached %d local file(s) to event media_urls",
                    self.name, len(file_attachments),
                )
        except Exception:
            logger.warning(
                "[%s] File content extraction failed (non-fatal), continuing",
                self.name, exc_info=True,
            )

        if not text and not media_urls:
            logger.debug("[%s] Empty message, skipping", self.name)
            return

        source = self.build_source(
            chat_id=chat_id,
            chat_name=getattr(message, "conversation_title", None),
            chat_type=chat_type,
            user_id=sender_id,
            user_name=sender_nick,
            user_id_alt=sender_staff_id if sender_staff_id else None,
        )

        # Parse timestamp
        create_at = getattr(message, "create_at", None)
        try:
            timestamp = (
                datetime.fromtimestamp(int(create_at) / 1000, tz=timezone.utc)
                if create_at
                else datetime.now(tz=timezone.utc)
            )
        except (ValueError, OSError, TypeError):
            timestamp = datetime.now(tz=timezone.utc)

        event = MessageEvent(
            text=text,
            message_type=msg_type,
            source=source,
            message_id=msg_id,
            raw_message=message,
            media_urls=media_urls,
            media_types=media_types,
            timestamp=timestamp,
        )

        logger.debug(
            "[%s] Message from %s in %s: %s",
            self.name,
            sender_nick,
            chat_id[:20] if chat_id else "?",
            text[:80] if text else "(media)",
        )
        await self.handle_message(event)

    @staticmethod
    def _extract_quoted_msg_text(container: Any, max_depth: int = 3) -> Optional[str]:
        """Extract a ``[引用] <body>`` string from a DingTalk reply container.

        When a user long-presses → quotes → replies in the DingTalk client,
        the inbound payload carries ``isReplyMsg=True`` + ``repliedMsg={...}``
        alongside the user's new text. The DingTalk Robot OpenAPI does not
        surface this through a typed SDK field, so the reference connector
        (dingtalk-openclaw-connector, ``core/message-handler.ts:163-240``)
        walks the raw dict. We mirror that logic here to preserve the
        quoted-context signal for the agent.

        ``max_depth`` bounds nested-quote recursion (matches openclaw's 3).
        Returns ``None`` when no quote is present or the body is empty.
        """
        if max_depth <= 0 or not container:
            return None
        if not container.get("isReplyMsg"):
            return None

        replied = container.get("repliedMsg")
        if not replied:
            return None

        msg_type = replied.get("msgType") or "text"

        raw_content = replied.get("content")
        if isinstance(raw_content, dict):
            content_obj = raw_content
        elif isinstance(raw_content, str):
            try:
                parsed = json.loads(raw_content)
                content_obj = parsed if isinstance(parsed, dict) else {}
            except (ValueError, TypeError):
                content_obj = {}
        else:
            content_obj = {}

        body_text = ""
        if msg_type == "text":
            body_text = (content_obj.get("text") or replied.get("text") or "").strip()
            if content_obj.get("isReplyMsg"):
                nested = DingTalkAdapter._extract_quoted_msg_text(content_obj, max_depth - 1)
                if nested:
                    body_text = f"{body_text}\n{nested}" if body_text else nested
        elif msg_type == "richText":
            rich_list = content_obj.get("richText") or []
            parts = [
                item.get("text", "")
                for item in rich_list
                if isinstance(item, dict)
                and item.get("text")
                and item.get("msgType") != "skill"
                and not item.get("skillData")
            ]
            body_text = "".join(parts)
        elif msg_type == "picture":
            body_text = "[图片]"
        elif msg_type == "video":
            body_text = "[视频]"
        elif msg_type == "audio":
            body_text = content_obj.get("recognition") or "[语音消息]"
        elif msg_type == "file":
            file_name = content_obj.get("fileName") or "unknown"
            body_text = f"[文件: {file_name}]"
        elif msg_type == "markdown":
            body_text = (content_obj.get("text") or "").strip() or "[markdown消息]"
        elif msg_type == "interactiveCard":
            # Forward-compatible extraction.  Today DingTalk does NOT populate
            # `content` for quoted interactiveCards — only 4 metadata fields
            # (msgId/senderId/msgType/createdAt) are delivered.  We are
            # pushing the DingTalk IM team to populate `content.text` /
            # `content.markdown` / `content.title` on the server side; once
            # they ship, hermes picks it up automatically without code change.
            card_text = ""
            for key in ("text", "markdown", "title", "summary"):
                candidate = content_obj.get(key)
                if isinstance(candidate, str) and candidate.strip():
                    card_text = candidate.strip()
                    break
            if card_text:
                body_text = card_text
            else:
                card_url = (
                    content_obj.get("biz_custom_action_url")
                    or replied.get("biz_custom_action_url")
                    or ""
                )
                body_text = (
                    f"收到交互式卡片链接：{card_url}"
                    if card_url
                    else "[interactiveCard消息]"
                )
        else:
            body_text = f"[{msg_type}消息]"

        if not body_text:
            return None
        return f"[引用] {body_text}"

    @staticmethod
    def _extract_text(message: "ChatbotMessage") -> str:
        """Extract plain text from a DingTalk chatbot message.

        Handles both legacy and current dingtalk-stream SDK payload shapes:
          * legacy: ``message.text`` was a dict ``{"content": "..."}``
          * >= 0.20: ``message.text`` is a ``TextContent`` dataclass whose
            ``__str__`` returns ``"TextContent(content=...)"`` — never fall
            back to ``str(text)`` without extracting ``.content`` first.
          * rich text moved from ``message.rich_text`` (list) to
            ``message.rich_text_content.rich_text_list`` (list of dicts).

        When the user quoted an earlier message (``isReplyMsg=True``), the
        quoted body is appended as ``[引用] <body>`` so the agent sees the
        conversational context, matching dingtalk-openclaw-connector.
        """
        text = getattr(message, "text", None) or ""

        # Handle TextContent object (SDK style). TextContent.from_dict routes
        # unknown JSON fields (isReplyMsg, repliedMsg) into ``.extensions``.
        quote_container: Optional[Dict[str, Any]] = None
        if hasattr(text, "content"):
            content = (text.content or "").strip()
            extensions = getattr(text, "extensions", None)
            if isinstance(extensions, dict) and extensions.get("isReplyMsg"):
                quote_container = extensions
        elif isinstance(text, dict):
            content = text.get("content", "").strip()
            if text.get("isReplyMsg"):
                quote_container = text
        else:
            content = str(text).strip()

        if quote_container:
            try:
                logger.info(
                    "dingtalk quote payload: %s",
                    json.dumps(quote_container, ensure_ascii=False, default=str),
                )
            except Exception:
                logger.info("dingtalk quote payload (repr): %r", quote_container)
            quoted = DingTalkAdapter._extract_quoted_msg_text(quote_container, max_depth=3)
            if quoted:
                content = f"{content}\n{quoted}" if content else quoted

        if not content:
            rich_text = getattr(message, "rich_text_content", None) or getattr(
                message, "rich_text", None
            )
            if rich_text:
                rich_list = getattr(rich_text, "rich_text_list", None) or rich_text
                if isinstance(rich_list, list):
                    parts = []
                    for item in rich_list:
                        if isinstance(item, dict):
                            t = item.get("text") or item.get("content") or ""
                            if t:
                                parts.append(t)
                        elif hasattr(item, "text") and item.text:
                            parts.append(item.text)
                    content = " ".join(parts).strip()

        # Do NOT strip "@bot" from the text.  The mention is a routing
        # signal (delivered structurally via callback `isInAtList`), and
        # regex-stripping @handles would collateral-damage e-mails
        # (alice@example.com), SSH URLs (git@github.com), and literal
        # references the user wrote ("what does @openai think").  Let the
        # LLM see the raw text — it handles "@bot hello" cleanly.

        # -----------------------------------------------------------------
        # For message types that the SDK does NOT parse into typed attrs
        # (audio / video / file), extract a reasonable text placeholder from
        # extensions['content'].  Matches connector's extractMessageContent.
        # -----------------------------------------------------------------
        if not content:
            msg_type_str = getattr(message, "message_type", "") or ""
            ext_content = (getattr(message, "extensions", {}) or {}).get("content", None)
            if isinstance(ext_content, str):
                try:
                    ext_content = json.loads(ext_content)
                except (ValueError, TypeError):
                    ext_content = None
            if isinstance(ext_content, dict):
                if msg_type_str == "audio":
                    content = (
                        ext_content.get("recognition")
                        or ext_content.get("recognition_text")
                        or "[语音消息]"
                    )
                elif msg_type_str == "video":
                    content = "[视频]"
                elif msg_type_str == "file":
                    fname = ext_content.get("fileName", "文件")
                    content = f"[文件: {fname}]"
                elif msg_type_str == "markdown":
                    content = ext_content.get("text", "").strip() or "[markdown消息]"

        return content

    def _extract_media(self, message: "ChatbotMessage"):
        """Extract media info from message. Returns (MessageType, [urls], [mime_types])."""
        msg_type = MessageType.TEXT
        media_urls = []
        media_types = []

        # Check for image/picture
        image_content = getattr(message, "image_content", None)
        if image_content:
            download_code = getattr(image_content, "download_code", None)
            if download_code:
                media_urls.append(download_code)
                media_types.append("image")
                msg_type = MessageType.PHOTO

        # Check for rich text with mixed content
        rich_text = getattr(message, "rich_text_content", None) or getattr(
            message, "rich_text", None
        )
        if rich_text:
            rich_list = getattr(rich_text, "rich_text_list", None) or rich_text
            if isinstance(rich_list, list):
                for item in rich_list:
                    if isinstance(item, dict):
                        dl_code = (
                            item.get("downloadCode") or item.get("download_code") or ""
                        )
                        item_type = item.get("type", "")
                        if dl_code:
                            mapped = DINGTALK_TYPE_MAPPING.get(item_type, "file")
                            media_urls.append(dl_code)
                            if mapped == "image":
                                media_types.append("image")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.PHOTO
                            elif mapped == "audio":
                                media_types.append("audio")
                                if msg_type == MessageType.TEXT:
                                    # DingTalk's "voice" rich-text item is a
                                    # native voice note — route through STT.
                                    # "audio" comes from file uploads only;
                                    # keep those as AUDIO (no auto-STT).
                                    if item_type == "voice":
                                        msg_type = MessageType.VOICE
                                    else:
                                        msg_type = MessageType.AUDIO
                            elif mapped == "video":
                                media_types.append("video")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.VIDEO
                            else:
                                media_types.append("application/octet-stream")
                                if msg_type == MessageType.TEXT:
                                    msg_type = MessageType.DOCUMENT

        msg_type_str = getattr(message, "message_type", "") or ""
        if msg_type_str == "picture" and not media_urls:
            msg_type = MessageType.PHOTO
        elif msg_type_str == "richText":
            msg_type = (
                MessageType.PHOTO
                if any("image" in t for t in media_types)
                else MessageType.TEXT
            )

        return msg_type, media_urls, media_types

    async def _download_images_to_local(
        self,
        media_urls: List[str],
        media_types: List[str],
        message: "ChatbotMessage",
    ) -> List[str]:
        """Download remote image URLs to local temp files.

        DingTalk's ``_resolve_media_codes`` replaces download codes with signed
        OSS URLs.  These signed URLs are **not universally accessible** — the
        OSS signature may be bound to specific request headers, IP ranges, or
        Referer, so the downstream ``vision_analyze_tool`` (which runs in a
        different network context) consistently fails with "Invalid image
        source" or HTTP 403.

        The reference connector (dingtalk-openclaw-connector) avoids this by
        downloading images to local files first and passing ``file://`` paths.
        We mirror that approach here: for each image in *media_urls*, download
        via httpx (which runs in the same process as the DingTalk SDK and
        therefore shares the same network/IP context) and replace the URL with
        the local temp-file path.

        Non-image entries and entries that fail to download are left unchanged.
        """
        result = list(media_urls)  # shallow copy
        robot_code = getattr(message, "robot_code", None) or self._client_id
        token: Optional[str] = None  # lazy-fetched

        for i, url in enumerate(media_urls):
            mtype = media_types[i] if i < len(media_types) else ""
            if mtype != "image":
                continue  # only process images

            try:
                download_url = url
                # If URL is still a download code (not http), resolve it first
                if not url.startswith("http"):
                    if token is None:
                        token = await self._get_access_token()
                    download_url = await self._resolve_single_download_url(
                        url, robot_code, token,
                    )
                    if not download_url:
                        logger.warning(
                            "[%s] Failed to resolve image download code, keeping original",
                            self.name,
                        )
                        continue

                # Download to persistent inbox (kept for agent tool re-access
                # long after _on_message returns).
                ext = ".png"  # safe default for images
                # Try to detect from URL path (before query params)
                url_path = download_url.split("?")[0]
                for img_ext in (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"):
                    if url_path.lower().endswith(img_ext):
                        ext = img_ext
                        break

                msg_id = getattr(message, "message_id", "") or ""
                local_path = await _download_file_to_inbox(
                    download_url, f"image{ext}", msg_id=msg_id, timeout=30.0,
                )
                if local_path:
                    result[i] = local_path
                    logger.info(
                        "[%s] Downloaded image to local: %s -> %s",
                        self.name, url[:60], local_path,
                    )
                else:
                    logger.warning(
                        "[%s] Image download failed, keeping original URL: %s",
                        self.name, url[:80],
                    )
            except Exception as exc:
                logger.warning(
                    "[%s] Image download error (non-fatal): %s", self.name, exc,
                )

        return result

    async def _extract_and_parse_file_attachments(
        self, message: "ChatbotMessage",
    ) -> Tuple[List[str], List[Tuple[str, str]]]:
        """Download rich-media attachments and describe them for the agent.

        Handles:
          * msgtype='file'  → download to inbox + parse if small text
          * msgtype='audio' → download + surface DingTalk STT recognition
          * msgtype='video' → download
          * richText items   → download each (non-picture) item

        DingTalk's SDK only parses ``text``/``picture``/``richText``; the raw
        ``content`` dict for ``file``/``audio``/``video`` lives in
        ``message.extensions['content']``.

        Returns ``(parts, attachments)``:
          * ``parts`` — text blocks injected into the user message so the
            LLM knows what was sent and where to find it.
          * ``attachments`` — ``[(local_path, mime_type), ...]`` the caller
            merges into ``event.media_urls`` / ``media_types`` so tools
            (``transcribe_audio``, ``vision_analyze``, ``read_file`` …) can
            open the file directly.

        Files live in ``~/.hermes/inbound_media/``; the cleanup sweep on
        adapter connect ages out anything older than 24 h.
        """
        parts: List[str] = []
        attachments: List[Tuple[str, str]] = []
        # (download_code_or_url, file_name, extra_text_hint)
        items: List[Tuple[str, str, str]] = []

        msg_type_str = getattr(message, "message_type", "") or ""
        extensions = getattr(message, "extensions", {}) or {}
        msg_id = getattr(message, "message_id", "") or ""

        raw_content: Optional[Dict[str, Any]] = None
        if msg_type_str in ("file", "audio", "video"):
            _raw = extensions.get("content", None)
            if isinstance(_raw, str):
                try:
                    _raw = json.loads(_raw)
                except (ValueError, TypeError):
                    _raw = None
            if isinstance(_raw, dict):
                raw_content = _raw

        # 1) file msgtype
        if msg_type_str == "file" and raw_content:
            dl = raw_content.get("downloadCode", "")
            fn = raw_content.get("fileName", "")
            if dl and fn:
                items.append((dl, fn, ""))
                logger.info("[%s] Found file attachment: %s", self.name, fn)

        # 2) audio msgtype — carry DingTalk's STT text as the extra hint
        if msg_type_str == "audio" and raw_content:
            dl = raw_content.get("downloadCode", "")
            fn = raw_content.get("fileName", "") or "audio.amr"
            recog = (
                raw_content.get("recognition")
                or raw_content.get("recognition_text")
                or ""
            )
            if dl:
                items.append((dl, fn, recog))
            elif recog:
                parts.append(
                    f"🎤 **音频**: {fn}\n"
                    f"📝 语音识别结果:\n{recog}"
                )

        # 3) video msgtype
        if msg_type_str == "video" and raw_content:
            dl = raw_content.get("downloadCode", "")
            fn = raw_content.get("fileName", "") or "video.mp4"
            if dl:
                items.append((dl, fn, ""))

        # 4) richText non-picture items (pictures handled upstream by
        #    ``_resolve_media_codes`` + ``_extract_media`` +
        #    ``_download_images_to_local``).
        rich_text = getattr(message, "rich_text_content", None)
        if rich_text:
            rich_list = getattr(rich_text, "rich_text_list", []) or []
            for item in rich_list:
                if not isinstance(item, dict):
                    continue
                dl = item.get("downloadCode") or item.get("download_code") or ""
                fn = item.get("fileName") or item.get("file_name") or ""
                itype = item.get("type", "")
                if dl and fn and itype not in ("picture",):
                    items.append((dl, fn, ""))

        # 5) Future-proofing: SDK typed ``file_content`` attribute
        if msg_type_str == "file" and not items:
            fc = getattr(message, "file_content", None)
            if fc:
                dl = getattr(fc, "download_code", None) or ""
                fn = getattr(fc, "file_name", None) or ""
                if dl and fn:
                    items.append((dl, fn, ""))

        if not items:
            return parts, attachments

        # Download every collected item into the persistent inbox and
        # describe it for the LLM.
        token = await self._get_access_token()
        robot_code = getattr(message, "robot_code", None) or self._client_id

        for dl_code, fname, extra in items:
            download_url = dl_code
            if not dl_code.startswith("http"):
                download_url = await self._resolve_single_download_url(
                    dl_code, robot_code, token,
                )
            if not download_url:
                parts.append(f"⚠️ 文件获取失败: {fname}")
                continue

            local_path = await _download_file_to_inbox(
                download_url, fname, msg_id=msg_id,
            )
            if not local_path:
                parts.append(f"⚠️ 文件下载失败: {fname}")
                continue

            mime = _mime_for_file(fname)
            attachments.append((local_path, mime))

            ext = os.path.splitext(fname)[1].lower()
            label = _file_type_label(ext)

            # Audio: STT hint + local path so the agent can re-transcribe.
            if extra:
                parts.append(
                    f"🎤 **{label}**: {fname}\n"
                    f"📝 语音识别结果:\n{extra}\n"
                    f"📁 本地路径: {local_path}"
                )
                continue

            # Text-parseable: inline if small, preview-only if large.
            if ext in _PARSEABLE_FILE_EXTS:
                try:
                    content = await asyncio.to_thread(
                        _parse_file_content, local_path, fname,
                    )
                except Exception:
                    content = None
                    logger.warning(
                        "[%s] parse_file_content raised for %s",
                        self.name, fname, exc_info=True,
                    )

                if content:
                    size = len(content)
                    logger.info(
                        "[%s] Parsed file %s: %d chars", self.name, fname, size,
                    )
                    if size <= 8000:
                        parts.append(
                            f"📄 **{label}**: {fname}（{size} 字符）\n"
                            f"📁 本地路径: {local_path}\n"
                            f"📋 内容:\n{content}"
                        )
                    else:
                        preview = content[:2000]
                        parts.append(
                            f"📄 **{label}**: {fname}（{size} 字符，已截断）\n"
                            f"📁 本地路径: {local_path}\n"
                            f"📋 前 2000 字预览:\n```\n{preview}\n```\n"
                            f"（完整内容请用 read_file 工具读取本地路径）"
                        )
                    continue
                # parse failed — fall through to path-only

            # Non-parseable / parse-failed: surface local path with tool hints.
            parts.append(
                f"📎 **{label}**: {fname}\n"
                f"📁 本地路径: {local_path}\n"
                f"（可用相应工具读取："
                f"transcribe_audio / vision_analyze / read_file）"
            )

        return parts, attachments

    async def _resolve_single_download_url(
        self, code: str, robot_code: str, token: Optional[str],
    ) -> Optional[str]:
        """Resolve a single downloadCode to a downloadUrl.

        Wraps the Robot SDK call without mutating any message attribute.
        """
        if not token or not self._robot_sdk:
            # Fallback: try direct HTTP with httpx
            if HTTPX_AVAILABLE and token:
                try:
                    async with httpx.AsyncClient(timeout=30.0) as client:
                        resp = await client.post(
                            "https://api.dingtalk.com/v1.0/robot/messageFiles/download",
                            json={"downloadCode": code, "robotCode": robot_code},
                            headers={
                                "x-acs-dingtalk-access-token": token,
                                "Content-Type": "application/json",
                            },
                        )
                    body = resp.json()
                    return body.get("downloadUrl")
                except Exception as exc:
                    logger.warning("[%s] httpx fallback download failed: %s", self.name, exc)
            return None
        try:
            request = dingtalk_robot_models.RobotMessageFileDownloadRequest(
                download_code=code,
                robot_code=robot_code,
            )
            headers = dingtalk_robot_models.RobotMessageFileDownloadHeaders(
                x_acs_dingtalk_access_token=token,
            )
            runtime = tea_util_models.RuntimeOptions()
            response = await self._robot_sdk.robot_message_file_download_with_options_async(
                request, headers, runtime,
            )
            body = response.body if response else None
            return getattr(body, "download_url", None) if body else None
        except Exception as exc:
            logger.warning("[%s] Failed to resolve download code: %s", self.name, exc)
            return None

    # -- Outbound messaging -------------------------------------------------

    async def send(
        self,
        chat_id: str,
        content: str,
        reply_to: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> SendResult:
        """Send a markdown reply via DingTalk session webhook.

        ``reply_to`` is accepted for cross-platform BasePlatformAdapter
        signature parity, but DingTalk Robot OpenAPI and session_webhook
        payloads have no "reply to message X" field.  The reference
        connector (dingtalk-openclaw-connector) also plumbs ``replyToId``
        through its signatures but never puts it in an outbound body —
        see messaging.ts:567-569.  We use it only as the "final reply"
        vs "intermediate tool progress" routing signal for AI Card
        lifecycle (see ``is_final_reply`` below).  Threading is likewise
        unsupported (see channel.ts:97 ``threads: false``).
        """
        metadata = metadata or {}
        logger.debug(
            "[%s] send() chat_id=%s card_enabled=%s",
            self.name,
            chat_id,
            bool(self._card_template_id and self._card_sdk),
        )

        # Check metadata first (for direct webhook sends)
        session_webhook = metadata.get("session_webhook")
        if not session_webhook:
            webhook_info = self._get_valid_webhook(chat_id)
            if webhook_info:
                session_webhook, _ = webhook_info
            else:
                # No reply webhook for this chat — this is a proactive send
                # (no preceding inbound message in cache).  Fall back to
                # Robot OpenAPI using AppKey/Secret, same pattern Feishu uses
                # for proactive ``im.v1.message.create`` calls.
                logger.debug(
                    "[%s] No session_webhook for chat_id=%s; routing via Robot OpenAPI",
                    self.name, chat_id,
                )
                result = await dingtalk_send_proactive(
                    client_id=self._client_id,
                    client_secret=self._client_secret,
                    robot_code=self._robot_code,
                    chat_id=chat_id,
                    content=content,
                )
                if result.get("success"):
                    return SendResult(
                        success=True,
                        message_id=result.get("request_id") or uuid.uuid4().hex[:12],
                    )
                return SendResult(
                    success=False, error=result.get("error", "unknown"),
                )

        if not self._http_client:
            return SendResult(success=False, error="HTTP client not initialized")

        # Look up the inbound message for this chat (for AI Card routing)
        current_message = self._message_contexts.get(chat_id)

        # ``reply_to`` is the signal that this send is the FINAL response
        # to an inbound user message — only `base.py:_send_with_retry` sets
        # it.  Tool-progress, commentary, and stream-consumer first-sends
        # all leave it None.  We use it for two orthogonal decisions:
        #   1. finalize on create?  Yes if final reply, No if intermediate
        #      (intermediate cards stay in streaming state so edit_message
        #      updates don't flicker closed→streaming→closed repeatedly).
        #   2. fire Done reaction?  Only when this is the final reply.
        is_final_reply = reply_to is not None

        # Try AI Card first (using alibabacloud_dingtalk.card_1_0 SDK).
        if self._card_template_id and current_message and self._card_sdk:
            # Close any previously-open streaming cards for this chat
            # before creating a new one (handles tool-progress → final-
            # response handoff; also cleans up lingering commentary cards).
            await self._close_streaming_siblings(chat_id)

            result = await self._create_and_stream_card(
                chat_id, current_message, content,
                finalize=is_final_reply,
            )
            if result and result.success:
                if is_final_reply:
                    # Final reply: card closed, swap Thinking → Done.
                    self._fire_done_reaction(chat_id)
                else:
                    # Intermediate (tool progress / commentary / streaming
                    # first chunk): keep the card open and track it so the
                    # next send() auto-closes it as a sibling, or
                    # edit_message(finalize=True) closes it explicitly.
                    self._streaming_cards.setdefault(chat_id, {})[
                        result.message_id
                    ] = content
                return result

            logger.warning("[%s] AI Card send failed, falling back to webhook", self.name)

        logger.debug("[%s] Sending via webhook", self.name)
        # Normalize markdown for DingTalk
        normalized = self._normalize_markdown(content[: self.MAX_MESSAGE_LENGTH])

        payload = {
            "msgtype": "markdown",
            "markdown": {"title": "Hermes", "text": normalized},
        }

        # Optional @mention: metadata may carry ``at_user_ids`` (list of
        # staff_ids) or ``at_all`` (bool) which DingTalk's custom-robot
        # webhook accepts under the ``at`` key.  Mirrors reference
        # messaging/send.ts:54-59.
        at_user_ids = metadata.get("at_user_ids") if metadata else None
        at_all = bool(metadata.get("at_all")) if metadata else False
        if at_user_ids or at_all:
            payload["at"] = {
                "atUserIds": list(at_user_ids) if at_user_ids else [],
                "isAtAll": at_all,
            }

        try:
            resp = await self._http_client.post(
                session_webhook, json=payload, timeout=15.0
            )
            if resp.status_code < 300:
                # Webhook path: fire Done only for final replies, same as
                # the card path.
                if is_final_reply:
                    self._fire_done_reaction(chat_id)
                return SendResult(success=True, message_id=uuid.uuid4().hex[:12])
            body = resp.text
            logger.warning(
                "[%s] Send failed HTTP %d: %s", self.name, resp.status_code, body[:200]
            )
            return SendResult(
                success=False, error=f"HTTP {resp.status_code}: {body[:200]}"
            )
        except httpx.TimeoutException:
            return SendResult(
                success=False, error="Timeout sending message to DingTalk"
            )
        except Exception as e:
            logger.error("[%s] Send error: %s", self.name, e)
            return SendResult(success=False, error=str(e))

    async def send_typing(self, chat_id: str, metadata=None) -> None:
        """DingTalk does not support typing indicators."""
        pass

    async def send_image(
        self,
        chat_id: str,
        image_url: str,
        caption: Optional[str] = None,
        reply_to: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> SendResult:
        """Send an image via DingTalk markdown.

        DingTalk's session webhook only supports text/markdown payloads, not
        native image/file attachments. For remote image URLs, render the image
        inline with markdown so the user still sees the image. Local files need
        OpenAPI media upload and are handled separately.
        """
        image_block = f"![image]({image_url})"
        content = f"{caption}\n\n{image_block}" if caption else image_block
        return await self.send(
            chat_id=chat_id,
            content=content,
            reply_to=reply_to,
            metadata=metadata,
        )

    async def send_image_file(
        self,
        chat_id: str,
        image_path: str,
        caption: Optional[str] = None,
        reply_to: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        **kwargs,
    ) -> SendResult:
        """DingTalk webhook replies cannot send local image files directly."""
        return SendResult(
            success=False,
            error=(
                "DingTalk session webhook replies do not support local image uploads. "
                "Only markdown/text replies are supported without OpenAPI media upload."
            ),
        )

    async def send_document(
        self,
        chat_id: str,
        file_path: str,
        caption: Optional[str] = None,
        file_name: Optional[str] = None,
        reply_to: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        **kwargs,
    ) -> SendResult:
        """DingTalk webhook replies cannot send local file attachments directly."""
        return SendResult(
            success=False,
            error=(
                "DingTalk session webhook replies do not support local file attachments. "
                "Only markdown/text replies are supported without OpenAPI message send."
            ),
        )

    async def get_chat_info(self, chat_id: str) -> Dict[str, Any]:
        """Return basic info about a DingTalk conversation."""
        return {
            "name": chat_id,
            "type": "group" if "group" in chat_id.lower() else "dm",
        }

    def _get_valid_webhook(self, chat_id: str) -> Optional[tuple[str, int]]:
        """Get a valid (non-expired) session webhook for the given chat_id."""
        info = self._session_webhooks.get(chat_id)
        if not info:
            return None
        webhook, expired_time_ms = info
        # Check expiry with 5-minute safety margin
        if expired_time_ms and expired_time_ms > 0:
            now_ms = int(datetime.now(tz=timezone.utc).timestamp() * 1000)
            safety_margin_ms = 5 * 60 * 1000
            if now_ms + safety_margin_ms >= expired_time_ms:
                # Expired, remove from cache
                self._session_webhooks.pop(chat_id, None)
                return None
        return info

    async def _create_and_stream_card(
        self,
        chat_id: str,
        message: Any,
        content: str,
        *,
        finalize: bool = True,
    ) -> Optional[SendResult]:
        """Create an AI Card, deliver it to the conversation, and stream initial content.

        Always called with ``finalize=True`` from ``send()`` (closed state).
        If the caller later issues ``edit_message(finalize=False)``, the
        DingTalk streaming_update API reopens the card into streaming
        state, and we track that in ``_streaming_cards`` for sibling
        cleanup on the next send.
        """
        try:
            token = await self._get_access_token()
            if not token:
                return None

            out_track_id = f"hermes_{uuid.uuid4().hex[:12]}"

            conversation_id = getattr(message, "conversation_id", "") or ""
            conversation_type = getattr(message, "conversation_type", "1")
            is_group = str(conversation_type) == "2"
            sender_staff_id = getattr(message, "sender_staff_id", "") or ""

            runtime = tea_util_models.RuntimeOptions()

            # Step 1: Create card with STREAM callback type
            create_request = dingtalk_card_models.CreateCardRequest(
                card_template_id=self._card_template_id,
                out_track_id=out_track_id,
                card_data=dingtalk_card_models.CreateCardRequestCardData(
                    card_param_map={"content": ""},
                ),
                callback_type="STREAM",
                im_group_open_space_model=(
                    dingtalk_card_models.CreateCardRequestImGroupOpenSpaceModel(
                        support_forward=True,
                    )
                ),
                im_robot_open_space_model=(
                    dingtalk_card_models.CreateCardRequestImRobotOpenSpaceModel(
                        support_forward=True,
                    )
                ),
            )

            create_headers = dingtalk_card_models.CreateCardHeaders(
                x_acs_dingtalk_access_token=token,
            )

            await self._card_sdk.create_card_with_options_async(
                create_request, create_headers, runtime
            )

            # Step 2: Deliver card to the conversation
            if is_group:
                open_space_id = f"dtv1.card//IM_GROUP.{conversation_id}"
                deliver_request = dingtalk_card_models.DeliverCardRequest(
                    out_track_id=out_track_id,
                    user_id_type=1,
                    open_space_id=open_space_id,
                    im_group_open_deliver_model=(
                        dingtalk_card_models.DeliverCardRequestImGroupOpenDeliverModel(
                            robot_code=self._robot_code,
                        )
                    ),
                )
            else:
                if not sender_staff_id:
                    logger.warning(
                        "[%s] AI Card skipped: missing sender_staff_id for DM",
                        self.name,
                    )
                    return None
                open_space_id = f"dtv1.card//IM_ROBOT.{sender_staff_id}"
                deliver_request = dingtalk_card_models.DeliverCardRequest(
                    out_track_id=out_track_id,
                    user_id_type=1,
                    open_space_id=open_space_id,
                    im_robot_open_deliver_model=(
                        dingtalk_card_models.DeliverCardRequestImRobotOpenDeliverModel(
                            space_type="IM_ROBOT",
                        )
                    ),
                )

            deliver_headers = dingtalk_card_models.DeliverCardHeaders(
                x_acs_dingtalk_access_token=token,
            )

            await self._card_sdk.deliver_card_with_options_async(
                deliver_request, deliver_headers, runtime
            )

            # Step 3: Stream initial content.  finalize=True closes the
            # card immediately (one-shot); finalize=False keeps it open
            # for streaming edit_message updates by out_track_id.
            await self._stream_card_content(
                out_track_id, token, content, finalize=finalize,
            )

            logger.info(
                "[%s] AI Card %s: %s",
                self.name,
                "created+finalized" if finalize else "created (streaming)",
                out_track_id,
            )
            return SendResult(success=True, message_id=out_track_id)

        except Exception as e:
            logger.warning(
                "[%s] AI Card create failed: %s\n%s",
                self.name, e, traceback.format_exc(),
            )
            return None

    async def edit_message(
        self,
        chat_id: str,
        message_id: str,
        content: str,
        *,
        finalize: bool = False,
    ) -> SendResult:
        """Edit an AI Card by streaming updated content.

        ``message_id`` is the out_track_id returned by the initial ``send()``
        call that created this card.  Callers (stream_consumer, tool
        progress) track their own ids independently so two parallel flows
        on the same chat_id don't interfere.
        """
        if not message_id:
            return SendResult(success=False, error="message_id required")

        # Throttle non-finalize edits per out_track_id.  DingTalk's
        # streaming_update endpoint 403s if two edits to the same card
        # land within ~500ms.  Finalize edits bypass — dropping them
        # would leave the card stuck in streaming state forever.
        # Optimistic update: stamp lastUpdate BEFORE the network call so
        # concurrent edits don't both pass the window check during
        # the await (mirrors reply-dispatcher.ts:607 pattern).
        now_ms = int(datetime.now(tz=timezone.utc).timestamp() * 1000)
        if not finalize:
            last_ms = self._card_last_edit_ms.get(message_id, 0)
            if now_ms - last_ms < _CARD_EDIT_THROTTLE_MS:
                logger.debug(
                    "[%s] edit_message throttled (%dms since last) for %s",
                    self.name, now_ms - last_ms, message_id,
                )
                return SendResult(success=True, message_id=message_id)
        self._card_last_edit_ms[message_id] = now_ms

        token = await self._get_access_token()
        if not token:
            return SendResult(success=False, error="No access token")

        try:
            await self._stream_card_content(
                message_id, token, content, finalize=finalize,
            )
            if finalize:
                # Remove from streaming-cards tracking and fire Done.  This
                # is the canonical "response ended" signal from stream
                # consumer's final edit.
                self._streaming_cards.get(chat_id, {}).pop(message_id, None)
                if not self._streaming_cards.get(chat_id):
                    self._streaming_cards.pop(chat_id, None)
                # Card is closed; throttle ts is no longer useful.
                self._card_last_edit_ms.pop(message_id, None)
                logger.debug(
                    "[%s] AI Card finalized (edit): %s",
                    self.name, message_id,
                )
                self._fire_done_reaction(chat_id)
            else:
                # Non-final edit reopens the card into streaming state —
                # track it so the next send() can auto-close it as a
                # sibling.
                self._streaming_cards.setdefault(chat_id, {})[message_id] = content
            return SendResult(success=True, message_id=message_id)
        except Exception as e:
            logger.warning("[%s] Card edit failed: %s", self.name, e)
            return SendResult(success=False, error=str(e))

    async def _stream_card_content(
        self,
        out_track_id: str,
        token: str,
        content: str,
        finalize: bool = False,
    ) -> None:
        """Stream content to an existing AI Card.

        Per-card 800ms throttle happens at the ``edit_message`` layer; this
        function additionally goes through the **global** token bucket so
        that many parallel chats can't collectively overrun the tenant-wide
        DingTalk card-API QPS cap (~40/s).  On an upstream limit response,
        ``trigger_backoff`` pauses every acquirer for 2s before spending
        the next token — same pattern as messaging/card.ts:93.
        """
        stream_request = dingtalk_card_models.StreamingUpdateRequest(
            out_track_id=out_track_id,
            guid=str(uuid.uuid4()),
            key="content",
            content=content[: self.MAX_MESSAGE_LENGTH],
            is_full=True,
            is_finalize=finalize,
            is_error=False,
        )

        stream_headers = dingtalk_card_models.StreamingUpdateHeaders(
            x_acs_dingtalk_access_token=token,
        )

        runtime = tea_util_models.RuntimeOptions()
        await _CARD_BUCKET.acquire()
        try:
            await self._card_sdk.streaming_update_with_options_async(
                stream_request, stream_headers, runtime
            )
        except Exception as e:
            # Detect DingTalk's 403 "QpsLimit" style responses.  The Tea SDK
            # raises with the error code embedded in the message — string
            # match is intentional (reference messaging/card.ts:107-125
            # does the same).
            err_msg = str(e)
            if "QpsLimit" in err_msg or "403" in err_msg or "qps" in err_msg.lower():
                logger.warning(
                    "[%s] Card QPS limit hit, backing off %dms: %s",
                    self.name, _CARD_API_QPS_BACKOFF_MS, err_msg[:160],
                )
                _CARD_BUCKET.trigger_backoff()
            raise

    async def _get_access_token(self) -> Optional[str]:
        """Get access token using SDK's cached token."""
        if not self._stream_client:
            return None
        try:
            # SDK's get_access_token is sync and uses requests
            token = await asyncio.to_thread(self._stream_client.get_access_token)
            return token
        except Exception as e:
            logger.error("[%s] Failed to get access token: %s", self.name, e)
            return None

    async def _send_emotion(
        self,
        open_msg_id: str,
        open_conversation_id: str,
        emoji_name: str,
        *,
        recall: bool = False,
    ) -> None:
        """Add or recall an emoji reaction on a message."""
        if not self._robot_sdk or not open_msg_id or not open_conversation_id:
            return
        action = "recall" if recall else "reply"
        try:
            token = await self._get_access_token()
            if not token:
                return

            emotion_kwargs = {
                "robot_code": self._robot_code,
                "open_msg_id": open_msg_id,
                "open_conversation_id": open_conversation_id,
                "emotion_type": 2,
                "emotion_name": emoji_name,
            }
            runtime = tea_util_models.RuntimeOptions()

            if recall:
                emotion_kwargs["text_emotion"] = (
                    dingtalk_robot_models.RobotRecallEmotionRequestTextEmotion(
                        emotion_id="2659900",
                        emotion_name=emoji_name,
                        text=emoji_name,
                        background_id="im_bg_1",
                    )
                )
                request = dingtalk_robot_models.RobotRecallEmotionRequest(
                    **emotion_kwargs,
                )
                sdk_headers = dingtalk_robot_models.RobotRecallEmotionHeaders(
                    x_acs_dingtalk_access_token=token,
                )
                await self._robot_sdk.robot_recall_emotion_with_options_async(
                    request, sdk_headers, runtime
                )
            else:
                emotion_kwargs["text_emotion"] = (
                    dingtalk_robot_models.RobotReplyEmotionRequestTextEmotion(
                        emotion_id="2659900",
                        emotion_name=emoji_name,
                        text=emoji_name,
                        background_id="im_bg_1",
                    )
                )
                request = dingtalk_robot_models.RobotReplyEmotionRequest(
                    **emotion_kwargs,
                )
                sdk_headers = dingtalk_robot_models.RobotReplyEmotionHeaders(
                    x_acs_dingtalk_access_token=token,
                )
                await self._robot_sdk.robot_reply_emotion_with_options_async(
                    request, sdk_headers, runtime
                )
            logger.info(
                "[%s] _send_emotion: %s %s on msg=%s",
                self.name, action, emoji_name, open_msg_id[:24],
            )
        except Exception:
            logger.debug(
                "[%s] _send_emotion %s failed", self.name, action, exc_info=True
            )

    async def _resolve_media_codes(self, message: "ChatbotMessage") -> None:
        """Resolve download codes in message to actual URLs."""
        token = await self._get_access_token()
        if not token:
            return

        robot_code = getattr(message, "robot_code", None) or self._client_id
        codes_to_resolve = []

        # Collect codes and references to update
        # 1. Single image content
        img_content = getattr(message, "image_content", None)
        if img_content and getattr(img_content, "download_code", None):
            codes_to_resolve.append((img_content, "download_code"))

        # 2. Rich text list
        rich_text = getattr(message, "rich_text_content", None)
        if rich_text:
            rich_list = getattr(rich_text, "rich_text_list", []) or []
            for item in rich_list:
                if isinstance(item, dict):
                    for key in ("downloadCode", "pictureDownloadCode", "download_code"):
                        if item.get(key):
                            codes_to_resolve.append((item, key))

        if not codes_to_resolve:
            return

        # Resolve all codes in parallel
        tasks = []
        for obj, key in codes_to_resolve:
            code = getattr(obj, key, None) if hasattr(obj, key) else obj.get(key)
            if code:
                tasks.append(
                    self._fetch_download_url(code, robot_code, token, obj, key)
                )

        await asyncio.gather(*tasks, return_exceptions=True)

    async def _fetch_download_url(
        self, code: str, robot_code: str, token: str, obj, key: str
    ) -> None:
        """Fetch download URL for a single code using the robot SDK."""
        if not self._robot_sdk:
            logger.warning(
                "[%s] Robot SDK not initialized, cannot resolve media code",
                self.name,
            )
            return
        try:
            request = dingtalk_robot_models.RobotMessageFileDownloadRequest(
                download_code=code,
                robot_code=robot_code,
            )
            headers = dingtalk_robot_models.RobotMessageFileDownloadHeaders(
                x_acs_dingtalk_access_token=token,
            )
            runtime = tea_util_models.RuntimeOptions()
            response = await self._robot_sdk.robot_message_file_download_with_options_async(
                request, headers, runtime
            )
            body = response.body if response else None
            if body:
                url = getattr(body, "download_url", None)
                if url:
                    if hasattr(obj, key):
                        setattr(obj, key, url)
                    elif isinstance(obj, dict):
                        obj[key] = url
            else:
                logger.warning(
                    "[%s] Failed to download media: empty response for code %s",
                    self.name,
                    code,
                )
        except Exception as e:
            logger.error("[%s] Error resolving media code %s: %s", self.name, code, e)

    @staticmethod
    def _normalize_markdown(text: str) -> str:
        """Normalize markdown for DingTalk's parser.

        DingTalk's markdown renderer has quirks:
        - Numbered lists need blank line before them
        - Indented code blocks may render incorrectly
        """
        lines = text.split("\n")
        out = []
        for i, line in enumerate(lines):
            # Ensure blank line before numbered list items
            is_numbered = re.match(r"^\d+\.\s", line.strip())
            if is_numbered and i > 0:
                prev = lines[i - 1]
                if prev.strip() and not re.match(r"^\d+\.\s", prev.strip()):
                    out.append("")
            # Dedent fenced code blocks
            if line.strip().startswith("```") and line != line.lstrip():
                indent = len(line) - len(line.lstrip())
                line = line[indent:]
            out.append(line)
        return "\n".join(out)


# ---------------------------------------------------------------------------
# Internal stream handler
# ---------------------------------------------------------------------------


class _IncomingHandler(
    dingtalk_stream.ChatbotHandler if DINGTALK_STREAM_AVAILABLE else object
):
    """dingtalk-stream ChatbotHandler that forwards messages to the adapter.

    SDK >= 0.20 changed process() from sync to async, and the message
    parameter from ChatbotMessage to CallbackMessage. We parse the
    CallbackMessage.data dict into a ChatbotMessage before forwarding.
    """

    def __init__(self, adapter: DingTalkAdapter, loop: Optional[asyncio.AbstractEventLoop] = None):
        if DINGTALK_STREAM_AVAILABLE:
            super().__init__()
        self._adapter = adapter
        self._loop = loop

    def pre_start(self) -> None:
        """No-op pre-start hook required by dingtalk-stream SDK.

        The SDK calls ``pre_start()`` on every registered handler before
        opening the WebSocket connection.  Without this method, the SDK
        raises ``AttributeError: '_IncomingHandler' object has no
        attribute 'pre_start'`` and kills the stream connection.
        """
        return

    async def process(self, message: "CallbackMessage"):
        """Called by dingtalk-stream (>=0.20) when a message arrives.

        dingtalk-stream >= 0.24 passes a CallbackMessage whose ``.data`` contains
        the chatbot payload. Convert it to ChatbotMessage via
        ``ChatbotMessage.from_dict()``.

        Message processing is dispatched as a background task so that this
        method returns the ACK immediately — blocking here would prevent the
        SDK from sending heartbeats, eventually causing a disconnect.
        """
        try:
            # CallbackMessage.data is a dict containing the raw DingTalk payload
            data = message.data
            if isinstance(data, str):
                data = json.loads(data)

            # Parse dict into ChatbotMessage using SDK's from_dict
            chatbot_msg = ChatbotMessage.from_dict(data)

            # Ensure session_webhook is populated even if the SDK's
            # from_dict() did not map it (field name mismatch across
            # SDK versions).
            if not getattr(chatbot_msg, "session_webhook", None):
                webhook = (
                    data.get("sessionWebhook")
                    or data.get("session_webhook")
                    or ""
                ) if isinstance(data, dict) else ""
                if webhook:
                    chatbot_msg.session_webhook = webhook

            # Ensure is_in_at_list is populated from the structured callback
            # flag even if from_dict() did not map it.  DingTalk sends
            # ``isInAtList`` in the raw payload; the adapter's mention check
            # reads the ChatbotMessage attribute ``is_in_at_list``.
            if not getattr(chatbot_msg, "is_in_at_list", False):
                raw_flag = (
                    data.get("isInAtList") if isinstance(data, dict) else False
                )
                if raw_flag:
                    chatbot_msg.is_in_at_list = True

            msg_id = getattr(chatbot_msg, "message_id", None) or ""
            conversation_id = getattr(chatbot_msg, "conversation_id", None) or ""

            # Thinking reaction — fire-and-forget, tracked
            if msg_id and conversation_id:
                self._adapter._spawn_bg(
                    self._adapter._send_emotion(
                        msg_id, conversation_id, "🤔Thinking", recall=False,
                    )
                )

            # Fire-and-forget: return ACK immediately, process in background.
            # Blocking here would prevent the SDK from sending heartbeats,
            # eventually causing a disconnect.  _on_message is wrapped so
            # exceptions inside the task surface in logs instead of
            # disappearing into the event loop.
            asyncio.create_task(self._safe_on_message(chatbot_msg))
        except Exception:
            logger.exception(
                "[%s] Error preparing incoming message", self._adapter.name
            )
            return AckMessage.STATUS_SYSTEM_EXCEPTION, "error"

        return AckMessage.STATUS_OK, "OK"

    async def _safe_on_message(self, chatbot_msg: "ChatbotMessage") -> None:
        """Wrapper that catches exceptions from _on_message.

        Dispatches through ``_enqueue_inbound`` so same-chat messages are
        serialized (with a busy-ACK on the tail), mirroring the reference
        connector's per-session queue at core/message-handler.ts:1597-1720.
        """
        try:
            await self._adapter._enqueue_inbound(chatbot_msg)
        except Exception:
            logger.exception(
                "[%s] Error processing incoming message", self._adapter.name
            )


# ──────────────────────────────────────────────────────────────────────────
# Plugin migration glue (#41112 / #3823)
#
# Added when the DingTalk adapter moved from gateway/platforms/dingtalk.py into
# this bundled plugin. Mirrors the Discord (#24356) / Slack migrations: a
# register(ctx) entry point plus hook implementations that replace the
# per-platform core touchpoints (the Platform.DINGTALK elif in gateway/run.py,
# the dingtalk_cfg YAML→env block + _PLATFORM_CONNECTED_CHECKERS entry in
# gateway/config.py, the _setup_dingtalk wizard + _PLATFORMS["dingtalk"] static
# dict in hermes_cli/gateway.py, and the _send_dingtalk dispatch in
# tools/send_message_tool.py).
# ──────────────────────────────────────────────────────────────────────────


async def _standalone_send(
    pconfig,
    chat_id,
    message,
    *,
    thread_id=None,
    media_files=None,
    force_document=False,
):
    """Out-of-process DingTalk delivery via a static robot webhook URL.

    Implements the standalone_sender_fn contract so deliver=dingtalk cron jobs
    succeed when cron runs separately from the gateway. The live adapter uses
    per-session webhook URLs from incoming messages, which aren't available
    out-of-process; this path uses the static DINGTALK_WEBHOOK_URL / extra
    webhook_url instead. Replaces the legacy _send_dingtalk helper.
    """
    extra = getattr(pconfig, "extra", {}) or {}
    try:
        import httpx
    except ImportError:
        return {"error": "httpx not installed"}
    try:
        webhook_url = extra.get("webhook_url") or os.getenv("DINGTALK_WEBHOOK_URL", "")
        if not webhook_url:
            return {"error": "DingTalk not configured. Set DINGTALK_WEBHOOK_URL env var or webhook_url in dingtalk platform extra config."}
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                webhook_url,
                json={"msgtype": "text", "text": {"content": message}},
            )
            resp.raise_for_status()
            data = resp.json()
            if data.get("errcode", 0) != 0:
                return {"error": f"DingTalk API error: {data.get('errmsg', 'unknown')}"}
        return {"success": True, "platform": "dingtalk", "chat_id": chat_id}
    except Exception as e:
        # Redact the access_token from webhook URLs that may appear in the
        # exception text. Reuse send_message_tool._error's redaction so the
        # logic stays single-sourced (lazy import avoids a circular at module
        # load). Falls back to a plain message if that helper is unavailable.
        try:
            from tools.send_message_tool import _error as _redact_error
            return _redact_error(f"DingTalk send failed: {e}")
        except Exception:
            return {"error": f"DingTalk send failed: {e}"}


def interactive_setup() -> None:
    """Configure DingTalk — QR scan (recommended) or manual credential entry.

    Replaces hermes_cli/setup.py-era _setup_dingtalk + the static
    _PLATFORMS["dingtalk"] dict in hermes_cli/gateway.py. CLI helpers are
    lazy-imported so the plugin's module-load surface stays minimal.
    """
    from hermes_cli.config import get_env_value, save_env_value
    from hermes_cli.setup import prompt_choice
    from hermes_cli.cli_output import (
        prompt,
        prompt_yes_no,
        print_header,
        print_success,
        print_warning,
    )

    print_header("DingTalk")
    existing = get_env_value("DINGTALK_CLIENT_ID")
    if existing:
        print_success(f"DingTalk is already configured (Client ID: {existing}).")
        if not prompt_yes_no("Reconfigure DingTalk?", False):
            return

    method = prompt_choice(
        "Choose setup method",
        [
            "QR Code Scan (Recommended, auto-obtain Client ID and Client Secret)",
            "Manual Input (Client ID and Client Secret)",
        ],
        default=0,
    )

    if method == 0:
        try:
            from hermes_cli.dingtalk_auth import dingtalk_qr_auth
        except ImportError as exc:
            print_warning(f"QR auth module failed to load ({exc}), falling back to manual input.")
            _manual_credential_entry(prompt, save_env_value, print_success)
            return
        result = dingtalk_qr_auth()
        if result is None:
            print_warning("QR auth incomplete, falling back to manual input.")
            _manual_credential_entry(prompt, save_env_value, print_success)
            return
        client_id, client_secret = result
        save_env_value("DINGTALK_CLIENT_ID", client_id)
        save_env_value("DINGTALK_CLIENT_SECRET", client_secret)
        print_success("DingTalk configured via QR scan!")
    else:
        _manual_credential_entry(prompt, save_env_value, print_success)


def _manual_credential_entry(prompt, save_env_value, print_success) -> None:
    client_id = prompt("DingTalk Client ID (app key)")
    if not client_id:
        return
    save_env_value("DINGTALK_CLIENT_ID", client_id)
    client_secret = prompt("DingTalk Client Secret", password=True)
    if client_secret:
        save_env_value("DINGTALK_CLIENT_SECRET", client_secret)
    print_success("DingTalk credentials saved")


def _apply_yaml_config(yaml_cfg: dict, dingtalk_cfg: dict) -> dict | None:
    """Translate config.yaml dingtalk: keys into DINGTALK_* env vars.

    Implements the apply_yaml_config_fn contract (#24849). Mirrors the legacy
    dingtalk_cfg block from gateway/config.py::load_gateway_config(). Env vars
    take precedence over YAML (each assignment guarded by not os.getenv(...)).
    Returns None — everything flows through env.
    """
    import json as _json
    if "require_mention" in dingtalk_cfg and not os.getenv("DINGTALK_REQUIRE_MENTION"):
        os.environ["DINGTALK_REQUIRE_MENTION"] = str(dingtalk_cfg["require_mention"]).lower()
    if "mention_patterns" in dingtalk_cfg and not os.getenv("DINGTALK_MENTION_PATTERNS"):
        os.environ["DINGTALK_MENTION_PATTERNS"] = _json.dumps(dingtalk_cfg["mention_patterns"])
    frc = dingtalk_cfg.get("free_response_chats")
    if frc is not None and not os.getenv("DINGTALK_FREE_RESPONSE_CHATS"):
        if isinstance(frc, list):
            frc = ",".join(str(v) for v in frc)
        os.environ["DINGTALK_FREE_RESPONSE_CHATS"] = str(frc)
    ac = dingtalk_cfg.get("allowed_chats")
    if ac is not None and not os.getenv("DINGTALK_ALLOWED_CHATS"):
        if isinstance(ac, list):
            ac = ",".join(str(v) for v in ac)
        os.environ["DINGTALK_ALLOWED_CHATS"] = str(ac)
    allowed = dingtalk_cfg.get("allowed_users")
    if allowed is not None and not os.getenv("DINGTALK_ALLOWED_USERS"):
        if isinstance(allowed, list):
            allowed = ",".join(str(v) for v in allowed)
        os.environ["DINGTALK_ALLOWED_USERS"] = str(allowed)
    return None


def _is_connected(config) -> bool:
    """DingTalk is connected when client_id + client_secret are present.

    Mirrors the legacy _PLATFORM_CONNECTED_CHECKERS[Platform.DINGTALK] entry.
    Reads from PlatformConfig.extra first, then env vars.
    """
    extra = getattr(config, "extra", {}) or {}
    return bool(
        (extra.get("client_id") or os.getenv("DINGTALK_CLIENT_ID"))
        and (extra.get("client_secret") or os.getenv("DINGTALK_CLIENT_SECRET"))
    )


def _build_adapter(config):
    """Factory wrapper that constructs DingTalkAdapter from a PlatformConfig."""
    return DingTalkAdapter(config)


def register(ctx) -> None:
    """Plugin entry point — called by the Hermes plugin system."""
    ctx.register_platform(
        name="dingtalk",
        label="DingTalk",
        adapter_factory=_build_adapter,
        check_fn=check_dingtalk_requirements,
        is_connected=_is_connected,
        validate_config=_is_connected,
        required_env=["DINGTALK_CLIENT_ID", "DINGTALK_CLIENT_SECRET"],
        install_hint="pip install 'dingtalk-stream>=0.20' httpx",
        setup_fn=interactive_setup,
        apply_yaml_config_fn=_apply_yaml_config,
        allowed_users_env="DINGTALK_ALLOWED_USERS",
        allow_all_env="DINGTALK_ALLOW_ALL_USERS",
        cron_deliver_env_var="DINGTALK_HOME_CHANNEL",
        standalone_sender_fn=_standalone_send,
        emoji="🐳",
        allow_update_command=True,
    )
