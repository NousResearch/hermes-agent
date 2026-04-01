#!/usr/bin/env python3
"""Set formulas in xlsx cells.

Usage:
    python apply_formula.py data.xlsx --cell B10 --formula '=SUM(B1:B9)'
    python apply_formula.py data.xlsx --cell C5 --formula '=AVERAGE(C1:C4)' --sheet Analysis
    python apply_formula.py data.xlsx --batch '[{"cell":"B10","formula":"=SUM(B1:B9)"}]'
"""

import argparse
import json
import sys


def apply_formula(path, cell, formula, sheet=None):
    """Set a formula in a specific cell of an xlsx file.

    Args:
        path: Path to the xlsx file.
        cell: Cell reference like 'B10'.
        formula: Excel formula string like '=SUM(B1:B9)'.
        sheet: Sheet name. Defaults to active sheet.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    ws[cell] = formula
    wb.save(path)


def apply_formulas_batch(path, formulas, sheet=None):
    """Apply multiple formulas at once.

    Args:
        path: Path to the xlsx file.
        formulas: List of dicts with 'cell' and 'formula' keys.
        sheet: Sheet name. Defaults to active sheet.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    for entry in formulas:
        ws[entry["cell"]] = entry["formula"]
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Set formulas in xlsx cells")
    parser.add_argument("path", help="Path to xlsx file")
    parser.add_argument("--cell", help="Cell reference (e.g., B10)")
    parser.add_argument("--formula", help="Excel formula (e.g., =SUM(B1:B9))")
    parser.add_argument("--sheet", default=None, help="Sheet name")
    parser.add_argument("--batch", default=None, help='JSON array of {"cell", "formula"} objects')
    args = parser.parse_args()

    if args.batch:
        formulas = json.loads(args.batch)
        apply_formulas_batch(args.path, formulas, sheet=args.sheet)
        print(f"Applied {len(formulas)} formulas to {args.path}")
    elif args.cell and args.formula:
        apply_formula(args.path, args.cell, args.formula, sheet=args.sheet)
        print(f"Set {args.cell} = {args.formula} in {args.path}")
    else:
        print("Error: provide --cell and --formula, or --batch", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
