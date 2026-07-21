#!/usr/bin/env python3
"""Embed charts into xlsx workbooks.

Usage:
    python chart.py data.xlsx --type bar --x Month --y Revenue
    python chart.py data.xlsx --type line --x Date --y Price --output chart.xlsx --title "Price Trend"
"""

import argparse
import os
import sys


# Map of chart type names to openpyxl chart classes
_CHART_TYPES = {
    "bar": "BarChart",
    "line": "LineChart",
    "pie": "PieChart",
    "scatter": "ScatterChart",
    "area": "AreaChart",
}


def create_chart(path, chart_type, x_col, y_col, output=None, title=None, sheet=None):
    """Embed a chart into an xlsx workbook.

    Args:
        path: Input xlsx file path.
        chart_type: One of 'bar', 'line', 'pie', 'scatter', 'area'.
        x_col: Column name for X axis / categories.
        y_col: Column name for Y axis / values.
        output: Optional output path. If None, modifies input file in place.
        title: Optional chart title.
        sheet: Sheet name to read data from.
    """
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference

    chart_classes = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }

    if chart_type not in chart_classes:
        raise ValueError(f"Unsupported chart type: {chart_type}. Use: {', '.join(chart_classes)}")

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb.active

    # Find column indices by header name
    headers = {cell.value: cell.column for cell in ws[1]}
    if x_col not in headers:
        raise ValueError(f"Column '{x_col}' not found. Available: {list(headers.keys())}")
    if y_col not in headers:
        raise ValueError(f"Column '{y_col}' not found. Available: {list(headers.keys())}")

    x_col_idx = headers[x_col]
    y_col_idx = headers[y_col]
    max_row = ws.max_row

    chart = chart_classes[chart_type]()
    if title:
        chart.title = title

    data_ref = Reference(ws, min_col=y_col_idx, min_row=1, max_row=max_row)
    cats_ref = Reference(ws, min_col=x_col_idx, min_row=2, max_row=max_row)

    if chart_type == "scatter":
        from openpyxl.chart import Series as ChartSeries

        x_values = Reference(ws, min_col=x_col_idx, min_row=2, max_row=max_row)
        y_values = Reference(ws, min_col=y_col_idx, min_row=2, max_row=max_row)
        series = ChartSeries(y_values, x_values, title=y_col)
        chart.series.append(series)
    else:
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

    chart.width = 18
    chart.height = 12

    # Add chart to a new sheet
    chart_sheet_name = "Chart"
    if chart_sheet_name in wb.sheetnames:
        del wb[chart_sheet_name]
    chart_ws = wb.create_sheet(chart_sheet_name)
    chart_ws.add_chart(chart, "A1")

    target = output or path
    wb.save(target)
    return target


def main():
    parser = argparse.ArgumentParser(description="Embed charts into xlsx workbooks")
    parser.add_argument("path", help="Input xlsx file path")
    parser.add_argument(
        "--type",
        required=True,
        dest="chart_type",
        choices=list(_CHART_TYPES.keys()),
        help="Chart type",
    )
    parser.add_argument("--x", required=True, dest="x_col", help="X axis / categories column name")
    parser.add_argument("--y", required=True, dest="y_col", help="Y axis / values column name")
    parser.add_argument("--output", default=None, help="Output file (default: modify input in place)")
    parser.add_argument("--title", default=None, help="Chart title")
    parser.add_argument("--sheet", default=None, help="Data sheet name")
    args = parser.parse_args()

    target = create_chart(
        args.path,
        chart_type=args.chart_type,
        x_col=args.x_col,
        y_col=args.y_col,
        output=args.output,
        title=args.title,
        sheet=args.sheet,
    )
    print(f"Chart created in {target}")


if __name__ == "__main__":
    main()
